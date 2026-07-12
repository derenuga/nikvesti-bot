"""
Версії бюджетного плану Миколаєва + рішення про зміни (задание №2, схема budget).

Горсовет вносить зміни в бюджет рішеннями сесії «викласти додаток у новій
редакції» — повною заміною, без явних дельт. До кожного рішення департамент
фінансів готує пакет із «Порівняльними таблицями»: xlsx, де на одному листі
бок о бок стоять «Чинна редакція» (лівий блок колонок) і «Запропонована
редакція з урахуванням змін» (правий блок), порядково вирівняні. Дельти
рахуємо як права мінус ліва. Саме ці xlsx — джерело завантаження.

UX: редакція просто кидає боту в приват ZIP пакета сесії (як він приходить
від депфіну) — без команд і аргументів. Бот сам розпаковує, знаходить
порівняльні таблиці, визначає рік/базове рішення з шапок і нашаровує ревізію.
Порядок нашаровування = порядок надсилання пакетів.

Уроки з реальних пакетів (нульовий s-fi-022 + зміни s-fi-001/003/005/007,
перевірено на файлах редакції 11.07.2026):
- структура тек щоразу інша («Порявняльні таблиці…», «Матеріали», «порівняльна
  таблиця», «Порівняльні таблиці додатків») → шукаємо xlsx по всьому архіву;
- ІМЕНА ФАЙЛІВ БРЕШУТЬ: у пакеті s-fi-003 файл «Додаток 5» містить таблицю до
  додатка 3, а «Додаток 1» і «Додаток 2» — байт-у-байт один файл → номер
  додатка беремо ТІЛЬКИ з шапки всередині xlsx, дублі відсіюємо по md5;
- у дохідній таблиці всередині є проміжні підсумки «Усього доходів (без
  урахування…)» БЕЗ коду → кінець таблиці лише рядок з кодом «Х»
  («Разом доходів» / «УСЬОГО»);
- коди доходів ієрархічні (10000000→11000000→11010100), групи лежать упереміш
  з листовими рядками → контроль сум по верхньому рівню (X0000000);
- пакет може не мати дохідної таблиці (квітень/липень міняли лише видатки) —
  ревізія успадковує доходи попередньої, це не помилка;
- нульовий пакет (вихідне рішення) — ПОВНІСТЮ PDF, xlsx немає → original
  створюється без рядків, заголовні цифри тягнемо з тексту рішення (він
  цифровий), а рядки доллються з лівої частини першої порівняльної таблиці
  (ТЗ: PDF-додатки не парсимо; на 2027 заплановано парсер PDF-додатків);
- зміни бувають і поза сесіями (розпорядження мера, виконком) — сума дельт
  рішень НЕ пояснює всю динаміку плану; місячні снапшоти (задание №1, ще не
  реалізоване; джерело — api.openbudget.gov.ua) — первинний факт, рішення —
  атрибуція. Снапшот на 01.07 = 7 077.9 млн vs наша редакція 30.04 =
  6 904.4 млн — різниця ~173.5 млн і є ці міжсесійні зміни.

Команди (додатково до «кинув zip»):
    /budget_load [рік] [номер] [дата] [base=50/26] — ручне завантаження одного
        xlsx (підпис до документа або reply); без аргументів усе визначається
        з шапки файлу, як у zip-флоу
    /budget_status — ревізії по роках, суми, розбіжності валідації
    /budget_headline <рік> <номер> ключ=значення … — поправити заголовні
        показники вручну (автоматом вони тягнуться з PDF рішення в пакеті)

Тихо не працює без BOT_DATABASE_URL — як archive_mirror/analytics_store.
"""

import asyncio
import hashlib
import os
import re
import zipfile
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from handlers import bot_db

_ALLOWED_USER_IDS = {
    int(uid)
    for uid in os.environ.get("ALLOWED_USER_IDS", "").split(",")
    if uid.strip()
}

# ---------- Схема (тільки нові таблиці, існуючих не чіпаємо) ----------

_BUDGET_SCHEMA_STATEMENTS = [
    "CREATE SCHEMA IF NOT EXISTS budget",
    # Реєстр редакцій плану: вихідне рішення + кожна зміна
    """
    CREATE TABLE IF NOT EXISTS budget.plan_revision (
        id              serial PRIMARY KEY,
        fiscal_year     int NOT NULL,
        decision_number text NOT NULL,
        decision_date   date,
        kind            text NOT NULL CHECK (kind IN ('original','amendment')),
        effective_order int NOT NULL,
        source_file     text,
        notes           text,
        UNIQUE (fiscal_year, effective_order)
    )
    """,
    # Рядки видаткової частини (структура Додатка 3)
    """
    CREATE TABLE IF NOT EXISTS budget.plan_expenditure_line (
        id                  bigserial PRIMARY KEY,
        revision_id         int NOT NULL REFERENCES budget.plan_revision,
        kpkvk               text NOT NULL,
        tpkvk               text,
        kfk                 text,
        line_name           text NOT NULL,
        is_unit_total       boolean DEFAULT false,
        general_total       numeric(16,2),
        general_consumption numeric(16,2),
        general_salary      numeric(16,2),
        general_utilities   numeric(16,2),
        general_development numeric(16,2),
        special_total       numeric(16,2),
        special_dev_budget  numeric(16,2),
        special_consumption numeric(16,2),
        special_salary      numeric(16,2),
        special_utilities   numeric(16,2),
        special_development numeric(16,2),
        total               numeric(16,2),
        UNIQUE (revision_id, kpkvk)
    )
    """,
    # Рядки дохідної частини (структура Додатка 1). Коди ієрархічні —
    # групи (X0000000 і проміжні) зберігаються нарівні з листовими рядками.
    """
    CREATE TABLE IF NOT EXISTS budget.plan_revenue_line (
        id                 bigserial PRIMARY KEY,
        revision_id        int NOT NULL REFERENCES budget.plan_revision,
        code               text NOT NULL,
        line_name          text NOT NULL,
        total              numeric(16,2),
        general_fund       numeric(16,2),
        special_fund       numeric(16,2),
        special_dev_budget numeric(16,2),
        UNIQUE (revision_id, code)
    )
    """,
    # Заголовні показники з текстової частини рішення (авто з PDF рішення
    # в пакеті + ручні поправки /budget_headline)
    """
    CREATE TABLE IF NOT EXISTS budget.plan_headline (
        revision_id           int PRIMARY KEY REFERENCES budget.plan_revision,
        revenue_total         numeric(16,2),
        revenue_general       numeric(16,2),
        revenue_special       numeric(16,2),
        expenditure_total     numeric(16,2),
        expenditure_general   numeric(16,2),
        expenditure_special   numeric(16,2),
        reserve_fund          numeric(16,2),
        debt_limit            numeric(16,2),
        guaranteed_debt_limit numeric(16,2),
        staff_count           int
    )
    """,
    # Розбіжності «чинної редакції» документа зі збереженою ревізією.
    # Розбіжності — реальність (одруківки депфіну), НЕ зупиняють load.
    """
    CREATE TABLE IF NOT EXISTS budget.revision_validation_issue (
        id             bigserial PRIMARY KEY,
        revision_id    int NOT NULL REFERENCES budget.plan_revision,
        table_kind     text NOT NULL,
        code           text NOT NULL,
        field          text NOT NULL,
        stored_value   numeric(16,2),
        document_value numeric(16,2),
        created_at     timestamptz NOT NULL DEFAULT now()
    )
    """,
    # Дельти між ревізіями (видатки). Попередник — ОСТАННЯ ревізія року з
    # видатковими рядками перед цією (не просто order-1: ревізія могла не мати
    # цієї таблиці, якщо рішення її не міняло). Підсумки розпорядників
    # виключено, щоб не задвоювати програми.
    """
    CREATE OR REPLACE VIEW budget.v_plan_amendments AS
    SELECT n.revision_id, r.decision_number, r.decision_date,
           n.kpkvk, n.line_name,
           n.total - COALESCE(p.total, 0)                 AS delta_total,
           n.general_total - COALESCE(p.general_total, 0) AS delta_general,
           n.special_total - COALESCE(p.special_total, 0) AS delta_special,
           (p.kpkvk IS NULL) AS is_new_program
    FROM budget.plan_expenditure_line n
    JOIN budget.plan_revision r ON r.id = n.revision_id AND r.kind = 'amendment'
    LEFT JOIN budget.plan_expenditure_line p
      ON p.kpkvk = n.kpkvk
     AND p.revision_id = (SELECT pr.id FROM budget.plan_revision pr
                          WHERE pr.fiscal_year = r.fiscal_year
                            AND (COALESCE(pr.decision_date, DATE '1900-01-01'), pr.effective_order)
                              < (COALESCE(r.decision_date, DATE '1900-01-01'), r.effective_order)
                            AND EXISTS (SELECT 1 FROM budget.plan_expenditure_line x
                                        WHERE x.revision_id = pr.id)
                          ORDER BY COALESCE(pr.decision_date, DATE '1900-01-01') DESC,
                                   pr.effective_order DESC LIMIT 1)
    WHERE NOT n.is_unit_total
      AND (n.total IS DISTINCT FROM p.total OR p.kpkvk IS NULL)
    """,
    # Дельти дохідної частини — той самий принцип
    """
    CREATE OR REPLACE VIEW budget.v_plan_revenue_amendments AS
    SELECT n.revision_id, r.decision_number, r.decision_date,
           n.code, n.line_name,
           n.total - COALESCE(p.total, 0)               AS delta_total,
           n.general_fund - COALESCE(p.general_fund, 0) AS delta_general,
           n.special_fund - COALESCE(p.special_fund, 0) AS delta_special,
           (p.code IS NULL) AS is_new_line
    FROM budget.plan_revenue_line n
    JOIN budget.plan_revision r ON r.id = n.revision_id AND r.kind = 'amendment'
    LEFT JOIN budget.plan_revenue_line p
      ON p.code = n.code
     AND p.revision_id = (SELECT pr.id FROM budget.plan_revision pr
                          WHERE pr.fiscal_year = r.fiscal_year
                            AND (COALESCE(pr.decision_date, DATE '1900-01-01'), pr.effective_order)
                              < (COALESCE(r.decision_date, DATE '1900-01-01'), r.effective_order)
                            AND EXISTS (SELECT 1 FROM budget.plan_revenue_line x
                                        WHERE x.revision_id = pr.id)
                          ORDER BY COALESCE(pr.decision_date, DATE '1900-01-01') DESC,
                                   pr.effective_order DESC LIMIT 1)
    WHERE n.total IS DISTINCT FROM p.total
       OR p.code IS NULL
    """,
    # Зниклі програми: рядок був у попередній видатковій редакції, у новій немає
    """
    CREATE OR REPLACE VIEW budget.v_plan_disappeared AS
    SELECT n.id AS revision_id, n.decision_number, n.decision_date,
           p.kpkvk, p.line_name, p.total AS last_total
    FROM budget.plan_revision n
    JOIN LATERAL (SELECT pr.id FROM budget.plan_revision pr
                  WHERE pr.fiscal_year = n.fiscal_year
                    AND (COALESCE(pr.decision_date, DATE '1900-01-01'), pr.effective_order)
                      < (COALESCE(n.decision_date, DATE '1900-01-01'), n.effective_order)
                    AND EXISTS (SELECT 1 FROM budget.plan_expenditure_line x
                                WHERE x.revision_id = pr.id)
                  ORDER BY COALESCE(pr.decision_date, DATE '1900-01-01') DESC,
                           pr.effective_order DESC LIMIT 1) prev ON true
    JOIN budget.plan_expenditure_line p ON p.revision_id = prev.id
    WHERE n.kind = 'amendment'
      AND NOT p.is_unit_total
      AND EXISTS (SELECT 1 FROM budget.plan_expenditure_line x
                  WHERE x.revision_id = n.id)
      AND NOT EXISTS (SELECT 1 FROM budget.plan_expenditure_line x
                      WHERE x.revision_id = n.id AND x.kpkvk = p.kpkvk)
    """,
]

_ensure_done = False


def ensure_budget_schema():
    """Ідемпотентно створює схему budget (лениво, раз на процес)."""
    global _ensure_done
    if _ensure_done:
        return
    for sql in _BUDGET_SCHEMA_STATEMENTS:
        bot_db.execute(sql)
    _ensure_done = True


def is_ready():
    return bot_db.is_configured()


# ---------- Парсер порівняльної таблиці ----------

# Скільки знаків у кодах: зеро-пад втрачених ведучих нулів
KPKVK_LEN = 7
TPKVK_LEN = 4
KFK_LEN = 4
REVENUE_CODE_LEN = 8

# Колонки блоку видатків (Додаток 3), позиції після рядка нумерації 1..16
_EXP_FIELDS = [
    "kpkvk", "tpkvk", "kfk", "line_name",
    "general_total", "general_consumption", "general_salary",
    "general_utilities", "general_development",
    "special_total", "special_dev_budget", "special_consumption",
    "special_salary", "special_utilities", "special_development",
    "total",
]
# Колонки блоку доходів (Додаток 1), нумерація 1..6
_REV_FIELDS = ["code", "line_name", "total", "general_fund", "special_fund", "special_dev_budget"]

_MONEY_FIELDS_EXP = _EXP_FIELDS[4:]
_MONEY_FIELDS_REV = _REV_FIELDS[2:]

# Верхній рівень ієрархії кодів доходів (10000000, 40000000…) — для контролю сум
_REV_TOP_RE = re.compile(r"^\d0{7}$")

_TITLE_YEAR_RE = re.compile(r"на\s+(20\d{2})\s+рік")
_TITLE_BASE_RE = re.compile(r"від\s+(\d{2}\.\d{2}\.\d{4})\s*№\s*([\d/]+)")
_TITLE_DODATOK_RE = re.compile(r"додатк\w*\s+(\d+)", re.I)
_SFI_RE = re.compile(r"s[\s_-]*fi[\s_-]*(\d+)", re.I)
# Токени з підпису до ZIP: дата ухвалення й ухвалений номер рішення (XX/YY)
_DATE_TOKEN = re.compile(r"\b(\d{2}\.\d{2}\.\d{4})\b")
_DECISION_TOKEN = re.compile(r"\b(\d{1,3}/\d{1,4})\b")


def _num_code(value, width):
    """210160.0 → '0210160'; 1.4549E9 → '1454900000'; 'Х'/порожнє → None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(int(round(value))).zfill(width)
    s = str(value).strip()
    if not s or s.upper() in ("Х", "X"):
        return None
    if re.fullmatch(r"\d+([.,]0+)?([eE]\+?\d+)?", s):
        try:
            return str(int(float(s.replace(",", ".")))).zfill(width)
        except ValueError:
            return None
    return None


def _is_x(value):
    return isinstance(value, str) and value.strip().upper() in ("Х", "X")


def _money(value):
    """Грошова клітинка → Decimal(2 знаки) або None (порожньо)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(round(value, 2)))
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s or s.upper() in ("Х", "X", "-"):
        return None
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _text(value):
    if value is None:
        return None
    s = re.sub(r"\s+", " ", str(value)).strip()
    return s or None


def _find_numbering_row(rows):
    """Шукає рядок нумерації колонок (1,2,3…). Повертає (індекс рядка,
    [(start_col, width), …] для кожного блока). Межа блоків — повторний старт
    нумерації з 1, НЕ фіксований номер колонки."""
    for i, row in enumerate(rows):
        starts = []  # (col, width)
        expect = None
        for col, v in enumerate(row):
            n = None
            if isinstance(v, (int, float)) and float(v).is_integer():
                n = int(v)
            elif isinstance(v, str) and v.strip().isdigit():
                n = int(v.strip())
            if n == 1:
                starts.append([col, 1])
                expect = 2
            elif expect and n == expect and starts:
                starts[-1][1] += 1
                expect += 1
            elif n is not None:
                expect = None
        # валідний рядок нумерації: 1..N (N≥5) один або два рази
        if starts and all(w >= 5 for _, w in starts):
            return i, [(c, w) for c, w in starts]
    return None, []


def _parse_block_row(row, start, width, kind):
    """Зріз рядка листа → dict полів рядка бюджету або None (порожньо/розділ/
    проміжний підсумок без коду)."""
    cells = list(row[start:start + width])
    cells += [None] * (width - len(cells))
    if kind == "expenditure":
        kpkvk = _num_code(cells[0], KPKVK_LEN)
        if not kpkvk:
            return None
        name = _text(cells[3])
        if not name:
            return None
        tpkvk = _num_code(cells[1], TPKVK_LEN)
        line = {
            "kpkvk": kpkvk,
            "tpkvk": tpkvk,
            "kfk": _num_code(cells[2], KFK_LEN),
            "line_name": name,
            # Підсумок розпорядника: КПКВК XX00000/XX10000 і немає ТПКВК
            "is_unit_total": tpkvk is None and kpkvk[-5:] in ("00000", "10000"),
        }
        for f, c in zip(_MONEY_FIELDS_EXP, cells[4:16]):
            line[f] = _money(c)
        return line
    else:
        code = _num_code(cells[0], REVENUE_CODE_LEN)
        name = _text(cells[1])
        if not code or not name:
            return None
        line = {"code": code, "line_name": name}
        for f, c in zip(_MONEY_FIELDS_REV, cells[2:6]):
            line[f] = _money(c)
        return line


def _grand_total_row(row, start, width, kind):
    """Фінальний підсумковий рядок — ТІЛЬКИ з кодом «Х» («УСЬОГО» видатків,
    «Разом доходів»). Проміжні «Усього…» без коду сюди не потрапляють —
    у дохідній таблиці вони стоять посеред даних."""
    cells = list(row[start:start + width])
    cells += [None] * (width - len(cells))
    if not _is_x(cells[0]):
        return None
    money_cells = cells[4:16] if kind == "expenditure" else cells[2:6]
    fields = _MONEY_FIELDS_EXP if kind == "expenditure" else _MONEY_FIELDS_REV
    return {f: _money(c) for f, c in zip(fields, money_cells)}


def parse_comparison_xlsx(data, filename=""):
    """Парсить xlsx порівняльної таблиці (або одноблочний додаток вихідного
    рішення). Повертає dict:
        kind        — 'expenditure' | 'revenue'
        dodatok     — номер додатка З ШАПКИ (імена файлів брешуть)
        blocks      — 1 (одноблочний оригінал) або 2 (порівняльна)
        left/right  — списки рядків (для одноблочного — тільки right)
        left_total/right_total — контрольні суми з рядка з кодом «Х»
        title/fiscal_year/base_number/base_date — з шапки
    Кидає ValueError з позначкою out_of_scope=True для додатків 2/4/5/6/7.
    """
    import openpyxl  # локальний імпорт: бот стартує і без openpyxl

    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    last_err = None
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        num_idx, blocks = _find_numbering_row(rows[:40])
        if num_idx is None:
            last_err = ValueError(f"лист '{sheet}': не знайшов рядок нумерації колонок")
            continue
        if len(blocks) not in (1, 2):
            last_err = ValueError(f"лист '{sheet}': {len(blocks)} блоків нумерації")
            continue

        title = re.sub(r"\s+", " ", " ".join(
            str(v) for row in rows[:num_idx] for v in row if isinstance(v, str)
        )).strip()

        # Номер додатка — тільки з шапки (у s-fi-003 файл «Додаток 5» містив
        # таблицю до додатка 3). У шапці два згадування («до додатку N до
        # рішення…») — беремо перше.
        m = _TITLE_DODATOK_RE.search(title)
        dodatok = int(m.group(1)) if m else None
        width = blocks[0][1]
        if dodatok is None:
            dodatok = 3 if width == 16 else (1 if width == 6 else None)
        if dodatok not in (1, 3):
            err = ValueError(f"додаток {dodatok} — поза схемою (вантажимо лише 1 і 3)")
            err.out_of_scope = True
            wb.close()
            raise err

        kind = "expenditure" if dodatok == 3 else "revenue"
        expected = 16 if kind == "expenditure" else 6
        if len(blocks) == 2 and blocks[1][1] != width:
            wb.close()
            raise ValueError(
                f"Блоки різної ширини: {blocks[0][1]} і {blocks[1][1]} — нова структура файлу?"
            )
        if width != expected:
            wb.close()
            raise ValueError(
                f"Додаток {dodatok}: ширина блоку {width}, очікував {expected} — "
                "структура змінилась, перевір файл"
            )

        left_rows, right_rows = [], []
        left_total = right_total = None
        for row in rows[num_idx + 1:]:
            gt = _grand_total_row(row, blocks[-1][0], width, kind)
            if gt is not None:
                right_total = gt
                if len(blocks) == 2:
                    left_total = _grand_total_row(row, blocks[0][0], width, kind)
                break
            right = _parse_block_row(row, blocks[-1][0], width, kind)
            left = _parse_block_row(row, blocks[0][0], width, kind) if len(blocks) == 2 else None
            if len(blocks) == 2:
                if left or right:
                    left_rows.append(left)
                    right_rows.append(right)
            elif right:
                right_rows.append(right)

        if not right_rows:
            last_err = ValueError(f"лист '{sheet}': не знайшов жодного рядка даних")
            continue

        ym = _TITLE_YEAR_RE.search(title)
        bm = _TITLE_BASE_RE.search(title)
        wb.close()
        return {
            "kind": kind,
            "dodatok": dodatok,
            "blocks": len(blocks),
            "left": left_rows,
            "right": right_rows,
            "left_total": left_total,
            "right_total": right_total,
            "filename": filename,
            "title": title,
            "fiscal_year": int(ym.group(1)) if ym else None,
            "base_date": datetime.strptime(bm.group(1), "%d.%m.%Y").date() if bm else None,
            "base_number": bm.group(2) if bm else None,
        }
    wb.close()
    raise ValueError(f"Не розібрав xlsx: {last_err}")


def _control_sum(lines, kind):
    """Сума для контролю проти рядка «Х»: видатки — програмні рядки без
    підсумків розпорядників; доходи — верхній рівень ієрархії (X0000000)."""
    if kind == "expenditure":
        return sum((ln["total"] or 0) for ln in lines if not ln["is_unit_total"])
    return sum((ln["total"] or 0) for ln in lines if _REV_TOP_RE.match(ln["code"]))


# ---------- Текст рішення (PDF у пакеті) → заголовні показники ----------

_HEADLINE_PATTERNS = [
    ("revenue_total", r"доходи бюджету[^;:]*?у сумі\s*(\d{6,})"),
    ("revenue_general", r"доходи загального фонду[^;:]*?[–—-]\s*(\d{6,})"),
    ("revenue_special", r"доходи спеціального фонду[^;:]*?[–—-]\s*(\d{6,})"),
    ("expenditure_total", r"видатки бюджету[^;:]*?у сумі\s*(\d{6,})"),
    ("expenditure_general", r"видатки загального фонду[^;:]*?[–—-]\s*(\d{6,})"),
    ("expenditure_special", r"видатки спеціального фонду[^;:]*?[–—-]\s*(\d{6,})"),
    ("reserve_fund", r"резервний фонд[^;:]*?розмірі\s*(\d{6,})"),
]


def parse_decision_pdf(data):
    """Витягає з цифрового PDF рішення заголовні показники і мету.
    Повертає dict (може бути частковим) або None, якщо текст не читається."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(BytesIO(data))
        text = " ".join((p.extract_text() or "") for p in reader.pages[:4])
    except Exception as e:  # noqa: BLE001 — PDF-скан або битий файл
        print(f"budget: PDF не читається — {e}")
        return None
    text = re.sub(r"\s+", " ", text)
    if not text.strip():
        return None
    out = {"is_original": "Про внесення змін" not in text[:600]}
    ym = _TITLE_YEAR_RE.search(text)
    if ym:
        out["fiscal_year"] = int(ym.group(1))
    bm = _TITLE_BASE_RE.search(text)
    if bm:
        out["base_date"] = datetime.strptime(bm.group(1), "%d.%m.%Y").date()
        out["base_number"] = bm.group(2)
    headline = {}
    for field, pat in _HEADLINE_PATTERNS:
        m = re.search(pat, text, re.I)
        if m:
            headline[field] = Decimal(m.group(1))
    m = re.search(r"чисельність[^;:.]*?(\d{2,6})\s*(?:штатн|одиниц)", text, re.I)
    if m:
        headline["staff_count"] = int(m.group(1))
    out["headline"] = headline
    return out


# ---------- Завантажувач ----------

_LINE_COLS = {
    "expenditure": ("budget.plan_expenditure_line",
                    ["kpkvk", "tpkvk", "kfk", "line_name", "is_unit_total"] + _MONEY_FIELDS_EXP),
    "revenue": ("budget.plan_revenue_line", ["code", "line_name"] + _MONEY_FIELDS_REV),
}


def _insert_lines(cur, revision_id, kind, lines):
    import psycopg2.extras
    table, cols = _LINE_COLS[kind]
    key = cols[0]
    # Ідемпотентна заміна: старі рядки цієї таблиці ревізії — геть, нові — на місце
    cur.execute(f"DELETE FROM {table} WHERE revision_id = %s", (revision_id,))
    seen, rows, dupes = set(), [], 0
    for ln in lines:
        if ln[key] in seen:  # дубль коду в документі — перший виграє
            dupes += 1
            continue
        seen.add(ln[key])
        rows.append(tuple([revision_id] + [ln.get(c) for c in cols]))
    psycopg2.extras.execute_values(
        cur,
        f"INSERT INTO {table} (revision_id, {', '.join(cols)}) VALUES %s",
        rows,
        page_size=200,
    )
    return len(rows), dupes


def _validate_left(cur, new_revision_id, kind, prev_revision_id, left_lines):
    """Звіряє «чинну редакцію» документа зі збереженою попередньою ревізією.
    Кожна розбіжність → рядок у revision_validation_issue. Нові програми
    (порожня ліва частина) сюди не потрапляють. Повертає к-сть розбіжностей."""
    table, cols = _LINE_COLS[kind]
    key = cols[0]
    money_fields = _MONEY_FIELDS_EXP if kind == "expenditure" else _MONEY_FIELDS_REV
    cur.execute(f"SELECT * FROM {table} WHERE revision_id = %s", (prev_revision_id,))
    colnames = [d[0] for d in cur.description]
    stored = {}
    for row in cur.fetchall():
        rec = dict(zip(colnames, row))
        stored[rec[key]] = rec

    issues = []
    doc_codes = set()
    for ln in left_lines:
        if ln is None:
            continue
        doc_codes.add(ln[key])
        st = stored.get(ln[key])
        if st is None:
            issues.append((ln[key], "presence", None, ln.get("total")))
            continue
        for f in money_fields:
            sv, dv = st.get(f), ln.get(f)
            if (sv or Decimal(0)) != (dv or Decimal(0)):
                issues.append((ln[key], f, sv, dv))
    for code, rec in stored.items():
        if code not in doc_codes:
            issues.append((code, "presence", rec.get("total"), None))

    for code, field, sv, dv in issues:
        cur.execute(
            "INSERT INTO budget.revision_validation_issue "
            "(revision_id, table_kind, code, field, stored_value, document_value) "
            "VALUES (%s, %s, %s, %s, %s, %s)",
            (new_revision_id, kind, code, field, sv, dv),
        )
    return len(issues)


def _check_unit_pairs(cur, revision_id, lines, block_label):
    """Внутрішня консистентність видаткового блока: пара «головний розпорядник
    (XX00000) ↔ відповідальний виконавець (XX10000)» має бути ідентичною.
    Розбіжність = одруківка депфіну (реальний кейс s-fi-001: 0200000/0210000
    мали різне розбиття споживання/розвиток при рівних підсумках) →
    revision_validation_issue з полем pair:<поле>. Повертає к-сть."""
    units = {ln["kpkvk"]: ln for ln in lines if ln and ln.get("is_unit_total")}
    found = 0
    for k, ln in units.items():
        if not (k.endswith("00000") and not k.endswith("10000")):
            continue
        mate = units.get(k[:2] + "10000")
        if not mate:
            continue
        for f in _MONEY_FIELDS_EXP:
            a, b = ln.get(f) or Decimal(0), mate.get(f) or Decimal(0)
            if a != b:
                cur.execute(
                    "INSERT INTO budget.revision_validation_issue "
                    "(revision_id, table_kind, code, field, stored_value, document_value) "
                    "VALUES (%s, 'expenditure', %s, %s, %s, %s)",
                    (revision_id, f"{k}|{k[:2]}10000", f"pair:{block_label}:{f}", a, b),
                )
                found += 1
    return found


def _pred_with_lines(cur, fiscal_year, order, kind, decision_date=None):
    """Ревізія року, що ХРОНОЛОГІЧНО передує цій, у якій є рядки цієї таблиці
    (ревізія могла не мати додатка 1/3, якщо рішення його не міняло).
    Хронологія — за датою ухвалення (щоб пакети можна було вантажити в будь-якому
    порядку); effective_order — лише тайбрейкер для ревізій без дати."""
    table, _ = _LINE_COLS[kind]
    cur.execute(
        f"""SELECT pr.id, pr.notes FROM budget.plan_revision pr
            WHERE pr.fiscal_year = %s
              AND (COALESCE(pr.decision_date, DATE '1900-01-01'), pr.effective_order)
                < (COALESCE(%s::date, DATE '1900-01-01'), %s)
              AND EXISTS (SELECT 1 FROM {table} x WHERE x.revision_id = pr.id)
            ORDER BY COALESCE(pr.decision_date, DATE '1900-01-01') DESC,
                     pr.effective_order DESC LIMIT 1""",
        (fiscal_year, decision_date, order),
    )
    return cur.fetchone()


def _get_or_create_revision(cur, fiscal_year, decision_number, decision_date,
                            kind_of_revision, filename, notes=None):
    """Знаходить ревізію (fiscal_year, decision_number) або створює нову з
    наступним effective_order. Повторне завантаження дубля не створює."""
    cur.execute(
        "SELECT id, effective_order FROM budget.plan_revision "
        "WHERE fiscal_year = %s AND decision_number = %s",
        (fiscal_year, decision_number),
    )
    row = cur.fetchone()
    if row:
        # повторне завантаження з датою (у першому проході її не було)
        if decision_date:
            cur.execute(
                "UPDATE budget.plan_revision SET decision_date = %s WHERE id = %s",
                (decision_date, row[0]),
            )
        return row[0], row[1], True
    if kind_of_revision == "original":
        # original року один: міг бути створений під іншим номером
        # (s-fi-код нульового пакета vs канонічний 50/26 з шапки таблиці)
        cur.execute(
            "SELECT id FROM budget.plan_revision "
            "WHERE fiscal_year = %s AND effective_order = 0",
            (fiscal_year,),
        )
        row = cur.fetchone()
        if row:
            return row[0], 0, True
        order = 0
    else:
        cur.execute(
            "SELECT COALESCE(MAX(effective_order), -1) + 1 "
            "FROM budget.plan_revision WHERE fiscal_year = %s",
            (fiscal_year,),
        )
        order = cur.fetchone()[0]
    cur.execute(
        "INSERT INTO budget.plan_revision "
        "(fiscal_year, decision_number, decision_date, kind, effective_order, source_file, notes) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id",
        (fiscal_year, decision_number, decision_date, kind_of_revision, order, filename, notes),
    )
    return cur.fetchone()[0], order, False


def load_parsed(parsed, fiscal_year, decision_number, decision_date=None,
                filename=""):
    """Завантажує один розібраний додаток (порівняльний або одноблочний
    оригінал) у схему budget. Синхронна — викликати через asyncio.to_thread."""
    ensure_budget_schema()
    kind = parsed["kind"]
    base_number = parsed.get("base_number")
    base_date = parsed.get("base_date")

    conn = bot_db._connect()
    try:
        with conn, conn.cursor() as cur:
            report = {
                "kind": kind, "blocks": parsed["blocks"],
                "original_created": False, "validated": 0, "issues": 0,
                "revision_reused": False, "dupes": 0,
            }
            right_present = [ln for ln in parsed["right"] if ln]

            if parsed["blocks"] == 1:
                # Одноблочний xlsx = додаток вихідного рішення → original
                revision_id, order, reused = _get_or_create_revision(
                    cur, fiscal_year, decision_number, decision_date,
                    "original", filename,
                )
                report["revision_reused"] = reused
                report["original_created"] = not reused
            else:
                cur.execute(
                    "SELECT count(*) FROM budget.plan_revision WHERE fiscal_year = %s",
                    (fiscal_year,),
                )
                have_revisions = cur.fetchone()[0] > 0
                left_present = [ln for ln in parsed["left"] if ln]

                if not have_revisions:
                    # Ревізій року ще немає: ліва частина стає original
                    # (номер/дата базового рішення — з шапки таблиці)
                    orig_id, _, _ = _get_or_create_revision(
                        cur, fiscal_year, base_number or "original", base_date,
                        "original", filename,
                        notes="reconstructed from comparison table",
                    )
                    _insert_lines(cur, orig_id, kind, left_present)
                    report["original_created"] = True
                elif base_number:
                    # Канонізуємо original: нульовий PDF-пакет міг створити його
                    # під кодом s-fi-XXX без дати — шапка порівняльної таблиці
                    # знає справжні номер і дату базового рішення
                    cur.execute(
                        "UPDATE budget.plan_revision "
                        "SET decision_number = %s, decision_date = COALESCE(decision_date, %s) "
                        "WHERE fiscal_year = %s AND effective_order = 0 "
                        "  AND kind = 'original' AND decision_number <> %s",
                        (base_number, base_date, fiscal_year, base_number),
                    )

                revision_id, order, reused = _get_or_create_revision(
                    cur, fiscal_year, decision_number, decision_date,
                    "amendment", filename,
                )
                report["revision_reused"] = reused

                # фактична дата ревізії (могла бути задана раніше /budget_date)
                cur.execute("SELECT decision_date FROM budget.plan_revision WHERE id = %s",
                            (revision_id,))
                eff_date = cur.fetchone()[0]
                pred = _pred_with_lines(cur, fiscal_year, order, kind, eff_date)
                if pred is None:
                    # Попередніх рядків цієї таблиці немає ніде: якщо original
                    # реконструйований/порожній (PDF-пакет) — доливаємо в нього
                    cur.execute(
                        "SELECT id, notes FROM budget.plan_revision "
                        "WHERE fiscal_year = %s AND effective_order = 0",
                        (fiscal_year,),
                    )
                    orig = cur.fetchone()
                    if orig and "reconstructed" in (orig[1] or ""):
                        _insert_lines(cur, orig[0], kind, left_present)
                        # original_created вимикає валідацію (це той самий файл),
                        # але pred лишаємо — щоб порахувати «було → стало»
                        report["original_created"] = True
                        pred = orig

                inserted, dupes = _insert_lines(cur, revision_id, kind, right_present)
                report.update(inserted=inserted, dupes=dupes)

                # повторне завантаження: старі issues цієї таблиці — геть
                cur.execute(
                    "DELETE FROM budget.revision_validation_issue "
                    "WHERE revision_id = %s AND table_kind = %s",
                    (revision_id, kind),
                )
                if pred is not None and not report["original_created"]:
                    report["issues"] = _validate_left(cur, revision_id, kind, pred[0], left_present)
                    report["validated"] = len(left_present)
                if kind == "expenditure":
                    report["pair_issues"] = (
                        _check_unit_pairs(cur, revision_id, left_present, "чинна")
                        + _check_unit_pairs(cur, revision_id, right_present, "нова")
                    )

                # Сума попередньої редакції — для «було → стало» у звіті
                if pred is not None:
                    table, _ = _LINE_COLS[kind]
                    flt = "AND NOT is_unit_total" if kind == "expenditure" else ""
                    if kind == "revenue":
                        flt = "AND code ~ '^\\d0{7}$'"
                    cur.execute(
                        f"SELECT COALESCE(SUM(total),0) FROM {table} "
                        f"WHERE revision_id = %s {flt}",
                        (pred[0],),
                    )
                    report["pred_total"] = cur.fetchone()[0]

            if parsed["blocks"] == 1:
                inserted, dupes = _insert_lines(cur, revision_id, kind, right_present)
                report.update(inserted=inserted, dupes=dupes)

            # Контроль сум: рядок «Х» документа vs сума завантаженого
            gt = parsed.get("right_total")
            if gt and gt.get("total") is not None:
                own = _control_sum(right_present, kind)
                report["doc_total"] = gt["total"]
                report["sum_total"] = own
                report["total_match"] = (own == gt["total"])
            report.update(revision_id=revision_id, effective_order=order)
            return report
    finally:
        conn.close()


def load_comparison(data, fiscal_year, decision_number, decision_date=None,
                    base_decision=None, filename=""):
    """Ручний шлях (/budget_load з одним xlsx): парсинг + завантаження.
    base_decision лишився для сумісності — база береться з шапки файлу."""
    parsed = parse_comparison_xlsx(data, filename=filename)
    if base_decision and not parsed.get("base_number"):
        parsed["base_number"] = base_decision
    return load_parsed(parsed, fiscal_year or parsed.get("fiscal_year"),
                       decision_number, decision_date, filename)


def upsert_headline(revision_id, values):
    """Записує заголовні показники ревізії (авто з PDF або вручну)."""
    if not values:
        return
    cols = list(values.keys())
    sets = ", ".join(f"{c} = COALESCE(EXCLUDED.{c}, budget.plan_headline.{c})" for c in cols)
    bot_db.execute(
        f"INSERT INTO budget.plan_headline (revision_id, {', '.join(cols)}) "
        f"VALUES (%s, {', '.join(['%s'] * len(cols))}) "
        f"ON CONFLICT (revision_id) DO UPDATE SET {sets}",
        tuple([revision_id] + [values[c] for c in cols]),
    )


# ---------- Пакет (ZIP або одиночний файл) ----------

def _zip_entries(data):
    """(ім'я, bytes) для файлів усередині zip. Імена лагодимо: мак-архіви
    пишуть UTF-8 без прапорця, і zipfile віддає cp437-кашу."""
    entries = []
    with zipfile.ZipFile(BytesIO(data)) as z:
        for info in z.infolist():
            if info.is_dir() or info.filename.startswith("__MACOSX"):
                continue
            name = info.filename
            if not (info.flag_bits & 0x800):
                try:
                    name = name.encode("cp437").decode("utf-8")
                except (UnicodeDecodeError, UnicodeEncodeError):
                    try:
                        name = name.encode("cp437").decode("cp866")
                    except (UnicodeDecodeError, UnicodeEncodeError):
                        pass
            entries.append((name, z.read(info)))
    return entries


def process_package(data, filename, decision_date=None, decision_override=None):
    """Обробляє пакет рішення: ZIP (або одиночний xlsx). Повертає dict:
        decision   — код рішення (s-fi-XXX або ім'я файлу)
        loads      — список звітів load_parsed по додатках 1/3
        skipped    — [(ім'я, причина)] пропущені файли
        headline   — dict з PDF рішення (або None)
        zero       — True, якщо пакет без порівняльних таблиць (нульовий PDF)
    decision_date/decision_override — з підпису до файлу: дата ухвалення й
    ухвалений номер рішення (у пакеті-проєкті їх немає — задаються вручну).
    Синхронна — викликати через asyncio.to_thread."""
    if filename.lower().endswith(".zip"):
        entries = _zip_entries(data)
    else:
        entries = [(filename, data)]

    m = _SFI_RE.search(filename) or next(
        (mm for name, _ in entries if (mm := _SFI_RE.search(name))), None
    )
    decision = decision_override or (
        f"s-fi-{int(m.group(1)):03d}" if m else re.sub(r"\.(zip|xlsx)$", "", filename, flags=re.I)
    )

    parsed_tables, skipped, seen_md5 = [], [], set()
    decision_pdf = None
    for name, blob in entries:
        low = name.lower()
        if low.endswith(".xlsx"):
            digest = hashlib.md5(blob).hexdigest()
            if digest in seen_md5:
                skipped.append((name, "дубль (той самий файл під іншим іменем)"))
                continue
            seen_md5.add(digest)
            try:
                parsed_tables.append(parse_comparison_xlsx(blob, filename=name))
            except ValueError as e:
                reason = "поза схемою (не додаток 1/3)" if getattr(e, "out_of_scope", False) else str(e)
                skipped.append((name, reason))
        elif low.endswith(".pdf"):
            # PDF самого рішення (не додаток і не пояснювальна) — заголовні цифри
            if ("додаток" not in low and "пояснювальна" not in low
                    and "записка" not in low and decision_pdf is None):
                decision_pdf = blob

    # Дубль таблиці до одного додатка (різні файли, однаковий вміст по суті)
    # — лишаємо перший
    by_dodatok = {}
    for p in parsed_tables:
        if p["dodatok"] in by_dodatok:
            skipped.append((p["filename"], f"другий файл додатка {p['dodatok']} — пропущено"))
        else:
            by_dodatok[p["dodatok"]] = p

    meta = parse_decision_pdf(decision_pdf) if decision_pdf else None
    fiscal_year = next(
        (p["fiscal_year"] for p in by_dodatok.values() if p["fiscal_year"]),
        (meta or {}).get("fiscal_year"),
    )
    if not fiscal_year:
        raise ValueError("Не визначив рік бюджету ні з шапок таблиць, ні з PDF рішення")

    result = {"decision": decision, "fiscal_year": fiscal_year, "loads": [],
              "skipped": skipped, "headline": None, "zero": False}

    if not by_dodatok:
        # Нульовий пакет: одні PDF. Створюємо original без рядків, заголовні
        # цифри — з тексту рішення. Рядки доллються з першої порівняльної.
        if not meta or not meta.get("is_original"):
            raise ValueError(
                "У пакеті немає порівняльних таблиць додатків 1/3, "
                "а PDF рішення не схожий на вихідний бюджет"
            )
        ensure_budget_schema()
        conn = bot_db._connect()
        try:
            with conn, conn.cursor() as cur:
                revision_id, order, reused = _get_or_create_revision(
                    cur, fiscal_year, decision, decision_date, "original", filename,
                    notes="original from PDF package; lines reconstructed from comparison table",
                )
        finally:
            conn.close()
        result["zero"] = True
        result["revision_id"] = revision_id
        result["revision_reused"] = reused
    else:
        # видатки після доходів — щоб у звіті «було → стало» по видатках
        # рахувалось уже з доліченим original
        for p in sorted(by_dodatok.values(), key=lambda x: x["dodatok"]):
            rep = load_parsed(p, fiscal_year, decision, decision_date, p["filename"])
            result["loads"].append(rep)
        result["revision_id"] = result["loads"][-1]["revision_id"]
        result["decision_date"] = decision_date

    if meta and meta.get("headline"):
        upsert_headline(result["revision_id"], meta["headline"])
        result["headline"] = meta["headline"]
    return result


# ---------- Аналітика для звіту в чат ----------

def amendments_summary(revision_id, kind, limit=8):
    """Найбільші дельти ревізії для повідомлення в чат."""
    view = "budget.v_plan_amendments" if kind == "expenditure" else "budget.v_plan_revenue_amendments"
    code_col = "kpkvk" if kind == "expenditure" else "code"
    new_col = "is_new_program" if kind == "expenditure" else "is_new_line"
    return bot_db.query(
        f"SELECT {code_col} AS code, line_name, delta_total, {new_col} AS is_new "
        f"FROM {view} WHERE revision_id = %s AND delta_total <> 0 "
        f"ORDER BY ABS(delta_total) DESC LIMIT %s",
        (revision_id, limit),
    )


# ---------- Telegram-хендлери ----------

def _fmt_money(v):
    if v is None:
        return "—"
    return f"{v:,.2f}".replace(",", " ").replace(".00", "")


def _parse_date(s):
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


_KIND_UA = {"expenditure": "Видатки (Додаток 3)", "revenue": "Доходи (Додаток 1)"}


def _package_reply(result):
    """Людський звіт по обробленому пакету."""
    d = result.get("decision_date")
    date_str = f" від {d.strftime('%d.%m.%Y')}" if d else ""
    lines = [f"📦 {result['decision']}{date_str} → бюджет {result['fiscal_year']}"]
    if not d and not result.get("zero"):
        lines.append("ℹ️ Дату ухвалення не задано (у пакеті-проєкті її немає). "
                     "Додай підписом до файлу «26.03.2026 51/47» або "
                     f"командою: /budget_date {result['fiscal_year']} {result['decision']} 26.03.2026")
    if result["zero"]:
        lines.append(
            "Це вихідне рішення (все в PDF): створив original-ревізію. Рядки "
            "додатків доллються з першої порівняльної таблиці, заголовні "
            "цифри взяв із тексту рішення."
        )
    for rep in result["loads"]:
        head = f"✅ {_KIND_UA[rep['kind']]}: ревізія #{rep['effective_order']}, {rep['inserted']} рядків"
        if "pred_total" in rep and rep.get("doc_total") is not None:
            head += f"\n     {_fmt_money(rep['pred_total'])} → {_fmt_money(rep['doc_total'])} грн"
        lines.append(head)
        if rep["original_created"]:
            lines.append("     📌 ліва частина збережена як original (перша таблиця року)")
        if rep["revision_reused"]:
            lines.append("     ♻️ повторне завантаження — без дубля ревізії")
        if rep.get("total_match") is False:
            lines.append(f"     ⚠️ контроль сум: документ {_fmt_money(rep['doc_total'])}, "
                         f"у мене {_fmt_money(rep['sum_total'])}")
        if rep["issues"]:
            lines.append(f"     ⚠️ «чинна редакція» розійшлась зі збереженою у {rep['issues']} "
                         f"місцях (міжсесійні зміни або одруківки) → revision_validation_issue")
        if rep.get("pair_issues"):
            lines.append(f"     ⚠️ одруківки депфіну: {rep['pair_issues']} розбіжностей у парах "
                         f"розпорядник/виконавець → revision_validation_issue")
        if rep.get("dupes"):
            lines.append(f"     ⚠️ дублі кодів у документі: {rep['dupes']}")
    if result.get("headline"):
        h = result["headline"]
        parts = []
        if h.get("revenue_total"):
            parts.append(f"доходи {_fmt_money(h['revenue_total'])}")
        if h.get("expenditure_total"):
            parts.append(f"видатки {_fmt_money(h['expenditure_total'])}")
        if parts:
            lines.append(f"📄 з тексту рішення: {', '.join(parts)} грн")
    import html
    for name, reason in result["skipped"]:
        short = html.escape(name.split("/")[-1])
        lines.append(f"⏭ {short} — {html.escape(reason)}")
    return lines


async def _load_top_deltas(lines, result):
    """Додає до звіту найбільші зміни по видатках (якщо це amendment) —
    з розпорядником, який освоюватиме гроші (КВК з перших цифр КПКВК)."""
    from handlers import budget_nlq
    for rep in result["loads"]:
        if rep["kind"] != "expenditure":
            continue
        is_amend = await bot_db.aquery(
            "SELECT 1 FROM budget.plan_revision WHERE id = %s AND kind = 'amendment'",
            (rep["revision_id"],),
        )
        if not is_amend:
            continue
        top = await asyncio.to_thread(amendments_summary, rep["revision_id"], "expenditure")
        if top:
            units = await asyncio.to_thread(budget_nlq._unit_names, result["fiscal_year"])
            lines.append("\n<b>Найбільші зміни у видатках:</b>")
            for t in top:
                lines.append(format_delta_block(
                    t["code"], t["line_name"], t["delta_total"],
                    units.get(t["code"][:2]), t["is_new"],
                ))
    return lines


def format_delta_block(code, name, delta, owner=None, is_new=False):
    """Читабельний блок однієї зміни для повідомлень (HTML-parse):
        🆕 <code> Назва
        +135 369 631 грн 🟢
        Розпорядник
    Порожній рядок між блоками робить довгі назви бюджетних програм
    придатними для читання (одним рядком вони зливаються)."""
    import html
    mark = "🆕 " if is_new else ""
    if delta is None:
        money = "— грн"
    else:
        sign = "+" if delta > 0 else ("−" if delta < 0 else "")
        dot = "🟢" if delta > 0 else ("🔴" if delta < 0 else "⚪")
        money = f"{sign}{_fmt_money(abs(delta))} грн {dot}"
    parts = [f"{mark}<b>{html.escape(str(code))}</b> {html.escape(str(name))}", money]
    if owner:
        parts.append(html.escape(str(owner)))
    return "\n" + "\n".join(parts)


async def budget_package_handler(update, context):
    """ZIP (або xlsx) пакета рішення в приваті бота — без команд і аргументів.
    Реагує лише на .zip/.xlsx від дозволених користувачів, решту ігнорує."""
    msg = update.effective_message
    doc = msg.document
    if not doc:
        return
    name = (doc.file_name or "").lower()
    if not (name.endswith(".zip") or name.endswith(".xlsx")):
        return
    if _ALLOWED_USER_IDS and update.effective_user.id not in _ALLOWED_USER_IDS:
        return
    if not is_ready():
        await msg.reply_text("Нора не налаштована (BOT_DATABASE_URL) — вантажити нікуди.")
        return

    f = await context.bot.get_file(doc.file_id)
    data = bytes(await f.download_as_bytearray())

    # Квартальний «Звіт про виконання … за N квартал» (ZIP казначейських форм
    # або окремий xlsx) — це виконання по КПКВК, окремий модуль
    from handlers import budget_execution_report as ber
    if ber.looks_like_execution_report(
        data=(None if name.endswith(".zip") else data),
        filename=doc.file_name,
    ):
        await ber.load_from_message(msg, context, data, doc.file_name)
        return

    # xlsx «Щомісячна інформація…» — це снапшот виконання (задание №1),
    # а не пакет рішення: роутимо в budget_snapshots
    if name.endswith(".xlsx"):
        from handlers import budget_snapshots
        if await asyncio.to_thread(budget_snapshots.looks_like_snapshot, data):
            await budget_snapshots.load_snapshot_from_message(msg, context, data, doc.file_name)
            return

    # Підпис до ZIP (опційно): дата ухвалення DD.MM.YYYY і/або ухвалений
    # номер рішення (напр. «26.03.2026 51/47») — у пакеті-проєкті їх немає
    caption = (msg.caption or "").strip()
    cap_date = _parse_date(_DATE_TOKEN.search(caption).group(0)) if _DATE_TOKEN.search(caption) else None
    cap_num = _DECISION_TOKEN.search(caption)
    cap_num = cap_num.group(0) if cap_num else None

    progress = await msg.reply_text(f"🦊 Розбираю {doc.file_name}…")
    try:
        result = await asyncio.to_thread(
            process_package, data, doc.file_name, cap_date, cap_num
        )
        lines = _package_reply(result)
        lines = await _load_top_deltas(lines, result)
    except Exception as e:  # noqa: BLE001 — повідомляємо в чат, не мовчимо
        await progress.edit_text(f"❌ Не завантажилось: {e}")
        return
    await progress.edit_text("\n".join(lines), parse_mode="HTML")


async def budget_load_handler(update, context):
    """/budget_load [рік] [номер] [дата] — ручний шлях для одного xlsx
    (підпис до документа або reply). Без аргументів усе визначається з файлу."""
    msg = update.effective_message
    if _ALLOWED_USER_IDS and update.effective_user.id not in _ALLOWED_USER_IDS:
        await msg.reply_text("⛔ Тільки для редакції.")
        return
    if not is_ready():
        await msg.reply_text("Нора не налаштована (BOT_DATABASE_URL) — вантажити нікуди.")
        return

    doc = msg.document or (msg.reply_to_message.document if msg.reply_to_message else None)
    fname = (doc.file_name or "") if doc else ""
    if not doc or not fname.lower().endswith((".xlsx", ".zip")):
        await msg.reply_text(
            "Просто надішли мені ZIP пакета рішення (або xlsx порівняльної таблиці) "
            "у приват — команда не потрібна. /budget_load лишився для ручних випадків: "
            "reply на файл із аргументами [рік] [номер] [дата]."
        )
        return

    args = (msg.caption or msg.text or "").split()[1:]
    plain = [a for a in args if not a.startswith("base=")]

    progress = await msg.reply_text(f"🦊 Розбираю {fname}…")
    try:
        f = await context.bot.get_file(doc.file_id)
        data = bytes(await f.download_as_bytearray())
        if fname.lower().endswith(".zip") or not plain:
            result = await asyncio.to_thread(process_package, data, fname)
        else:
            fiscal_year = int(plain[0]) if plain else None
            decision_number = plain[1] if len(plain) > 1 else fname
            decision_date = _parse_date(plain[2]) if len(plain) > 2 else None
            parsed = await asyncio.to_thread(parse_comparison_xlsx, data, fname)
            rep = await asyncio.to_thread(
                load_parsed, parsed, fiscal_year or parsed.get("fiscal_year"),
                decision_number, decision_date, fname,
            )
            result = {"decision": decision_number,
                      "fiscal_year": fiscal_year or parsed.get("fiscal_year"),
                      "loads": [rep], "skipped": [], "headline": None, "zero": False,
                      "revision_id": rep["revision_id"]}
        lines = _package_reply(result)
        lines = await _load_top_deltas(lines, result)
    except Exception as e:  # noqa: BLE001
        await progress.edit_text(f"❌ Не завантажилось: {e}")
        return
    await progress.edit_text("\n".join(lines), parse_mode="HTML")


async def budget_status_handler(update, context):
    """/budget_status — ревізії по роках, суми, розбіжності валідації."""
    msg = update.effective_message
    if not is_ready():
        await msg.reply_text("Нора не налаштована (BOT_DATABASE_URL).")
        return
    try:
        await asyncio.to_thread(ensure_budget_schema)
        revs = await bot_db.aquery(
            """
            SELECT r.id, r.fiscal_year, r.decision_number, r.decision_date, r.kind,
                   r.effective_order, r.notes,
                   (SELECT COUNT(*) FROM budget.plan_expenditure_line e
                     WHERE e.revision_id = r.id) AS exp_lines,
                   (SELECT COALESCE(SUM(e.total),0) FROM budget.plan_expenditure_line e
                     WHERE e.revision_id = r.id AND NOT e.is_unit_total) AS exp_total,
                   (SELECT COUNT(*) FROM budget.plan_revenue_line v
                     WHERE v.revision_id = r.id) AS rev_lines,
                   (SELECT COALESCE(SUM(v.total),0) FROM budget.plan_revenue_line v
                     WHERE v.revision_id = r.id AND v.code ~ '^\\d0{7}$') AS rev_total,
                   (SELECT COUNT(*) FROM budget.revision_validation_issue i
                     WHERE i.revision_id = r.id) AS issues
            FROM budget.plan_revision r
            ORDER BY r.fiscal_year, COALESCE(r.decision_date, DATE '1900-01-01'), r.effective_order
            """
        )
    except Exception as e:  # noqa: BLE001
        await msg.reply_text(f"❌ {e}")
        return
    if not revs:
        await msg.reply_text(
            "Схема budget порожня. Кинь мені ZIP пакета рішення (нульовий бюджет "
            "або зміни) у приват — розберу сам."
        )
        return
    lines, year = ["🦊 Бюджет у норі:"], None
    for r in revs:
        if r["fiscal_year"] != year:
            year = r["fiscal_year"]
            lines.append(f"\n📅 {year}:")
        d = f" від {r['decision_date'].strftime('%d.%m.%Y')}" if r["decision_date"] else ""
        parts = []
        if r["exp_lines"]:
            parts.append(f"видатки {r['exp_lines']} рядків, {_fmt_money(r['exp_total'])} грн")
        else:
            parts.append("видатки: успадковані" if r["kind"] == "amendment" else "видатки: рядків ще немає")
        if r["rev_lines"]:
            parts.append(f"доходи {_fmt_money(r['rev_total'])} грн")
        elif r["kind"] == "amendment":
            parts.append("доходи: успадковані")
        if r["issues"]:
            parts.append(f"⚠️ {r['issues']} розбіжностей")
        lines.append(
            f"#{r['effective_order']} {r['kind']} {r['decision_number']}{d} — " + "; ".join(parts)
        )
    lines.append(
        "\nℹ️ Виконання по розпорядниках і звірка з місячними снапшотами — "
        "/budget_execution (різниця з планом знімка = міжсесійні зміни)."
    )
    await msg.reply_text("\n".join(lines))


async def budget_date_handler(update, context):
    """/budget_date <рік> <код> <DD.MM.YYYY> [новий_номер] — задати дату
    ухвалення ревізії (у пакеті-проєкті її немає) і, опційно, канонічний
    номер рішення замість коду проєкту (s-fi-003 → 51/47)."""
    msg = update.effective_message
    if _ALLOWED_USER_IDS and update.effective_user.id not in _ALLOWED_USER_IDS:
        await msg.reply_text("⛔ Тільки для редакції.")
        return
    if not is_ready():
        await msg.reply_text("Нора не налаштована (BOT_DATABASE_URL).")
        return
    args = context.args or []
    if len(args) < 3:
        await msg.reply_text(
            "Формат: /budget_date 2026 s-fi-003 26.03.2026 [51/47]\n"
            "(дата ухвалення рішення; опційно — ухвалений номер замість коду проєкту)"
        )
        return
    try:
        fiscal_year = int(args[0])
    except ValueError:
        await msg.reply_text(f"Рік не число: {args[0]}")
        return
    code = args[1]
    d = _parse_date(args[2])
    if not d:
        await msg.reply_text(f"Дата не розпізнана (треба DD.MM.YYYY): {args[2]}")
        return
    new_number = args[3] if len(args) > 3 else None
    try:
        await asyncio.to_thread(ensure_budget_schema)
        rev = await bot_db.aquery(
            "SELECT id FROM budget.plan_revision WHERE fiscal_year = %s AND decision_number = %s",
            (fiscal_year, code),
        )
        if not rev:
            await msg.reply_text(f"Ревізії {code}/{fiscal_year} немає — спочатку завантаж пакет.")
            return
        await asyncio.to_thread(
            bot_db.execute,
            "UPDATE budget.plan_revision SET decision_date = %s"
            + (", decision_number = %s" if new_number else "")
            + " WHERE id = %s",
            ((d, new_number, rev[0]["id"]) if new_number else (d, rev[0]["id"])),
        )
    except Exception as e:  # noqa: BLE001
        await msg.reply_text(f"❌ {e}")
        return
    tail = f", номер → {new_number}" if new_number else ""
    await msg.reply_text(f"✅ {code}/{fiscal_year}: дата ухвалення {d.strftime('%d.%m.%Y')}{tail}")


_HEADLINE_FIELDS = {
    "revenue_total", "revenue_general", "revenue_special",
    "expenditure_total", "expenditure_general", "expenditure_special",
    "reserve_fund", "debt_limit", "guaranteed_debt_limit", "staff_count",
}


async def budget_headline_handler(update, context):
    """/budget_headline <рік> <номер> ключ=значення … — поправити заголовні
    показники вручну (автоматом тягнуться з PDF рішення при завантаженні)."""
    msg = update.effective_message
    if _ALLOWED_USER_IDS and update.effective_user.id not in _ALLOWED_USER_IDS:
        await msg.reply_text("⛔ Тільки для редакції.")
        return
    if not is_ready():
        await msg.reply_text("Нора не налаштована (BOT_DATABASE_URL).")
        return
    args = context.args or []
    if len(args) < 3:
        await msg.reply_text(
            "Формат: /budget_headline 2026 s-fi-001 revenue_total=6370506270 …\n"
            "Ключі: " + ", ".join(sorted(_HEADLINE_FIELDS))
        )
        return
    try:
        fiscal_year = int(args[0])
    except ValueError:
        await msg.reply_text(f"Рік не число: {args[0]}")
        return
    decision_number = args[1]
    values = {}
    for pair in args[2:]:
        if "=" not in pair:
            await msg.reply_text(f"Не ключ=значення: {pair}")
            return
        k, v = pair.split("=", 1)
        if k not in _HEADLINE_FIELDS:
            await msg.reply_text(f"Невідомий ключ: {k}")
            return
        try:
            values[k] = int(v) if k == "staff_count" else Decimal(v.replace(" ", ""))
        except (ValueError, InvalidOperation):
            await msg.reply_text(f"Не число: {pair}")
            return
    try:
        await asyncio.to_thread(ensure_budget_schema)
        rev = await bot_db.aquery(
            "SELECT id FROM budget.plan_revision WHERE fiscal_year = %s AND decision_number = %s",
            (fiscal_year, decision_number),
        )
        if not rev:
            await msg.reply_text(
                f"Ревізії {decision_number}/{fiscal_year} немає — спочатку завантаж пакет."
            )
            return
        await asyncio.to_thread(upsert_headline, rev[0]["id"], values)
    except Exception as e:  # noqa: BLE001
        await msg.reply_text(f"❌ {e}")
        return
    await msg.reply_text(
        f"✅ Заголовні показники {decision_number}/{fiscal_year}: "
        + ", ".join(f"{k}={_fmt_money(v)}" for k, v in values.items())
    )
