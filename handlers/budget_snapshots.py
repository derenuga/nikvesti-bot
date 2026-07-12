"""
Місячні снапшоти виконання бюджету Миколаєва (задание №1, схема budget).

Джерело — «Щомісячна інформація» на сайті міськради (публікується на початку
місяця, станом на 1-ше число):
    https://mkrada.gov.ua/content/shchomisyachna-informaciya-{РІК}r.html
Пари файлів за місяць: «Витрати щомісячна інформація DD.MM.YYYY.xlsx» і
«Надходження щомісячна інформація DD.MM.YYYY.xlsx».

Чому сайт міськради, а не OpenBudget API: казначейське API
(api.openbudget.gov.ua/api/public/localBudgetData) віддає видатки
консолідовано по ТПКВК — БЕЗ розпорядників. А редакції цінний саме розріз
«який розпорядник (чиновник) не виконав план у термін» — його дають лише
міськрадівські xlsx: видатки по КВК (розпорядниках) × КЕКВ (Лист1) і
× КБП-функціях (Лист2), з колонками «План на рік з урахуванням змін /
План за вказаний період / Касові видатки / % виконання». Доходи — за
назвами показників по фондах. OpenBudget лишається незалежною звіркою
підсумків (див. docs/BUDGET_REVISIONS_MODULE.md).

Методологія: снапшот витрат «за винятком надання та повернення кредитів та
БЕЗ видатків за рахунок власних надходжень бюджетних установ» — тому його
«План на рік зі змінами» НЕ дорівнює сумі ревізії рішень (budget_revisions)
один в один; різниця = міжсесійні зміни ± методологія. Це аналітичний
сигнал, а не помилка — звірка в /budget_execution подається з застереженням.

Дата зрізу — З ІМЕНІ ФАЙЛУ (у шапках xlsx її немає). Фіскальний рік =
(дата зрізу − 1 день).рік: файл за 01.01.2027 — це підсумок 2026-го.

Перший запуск монітора: бекфілимо в нору ВСЮ доступну історію року тихо,
звіт у чат шлемо лише за останній місяць (N=1 місяць = 2 файли) — щоб
перевірити формат відправки без спаму за 7 місяців (правило baseline
з CLAUDE.md).

Команди:
    /budget_execution      — виконання останнього снапшота по розпорядниках
                             (найгірші %, звірка з ревізією рішень)
    /budget_snapshot_check — перевірити сторінку міськради зараз (нові файли)
Файл «Щомісячна інформація…xlsx» можна кинути боту в приват і руками —
роутиться сюди з budget_package_handler.

Тихо не працює без BOT_DATABASE_URL — як budget_revisions.
"""

import asyncio
import json
import os
import re
import urllib.parse
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

import requests

from handlers import bot_db
from handlers.notifier import notify_error

MKRADA_BASE = "https://mkrada.gov.ua"
PAGE_URL_TMPL = MKRADA_BASE + "/content/shchomisyachna-informaciya-{year}r.html"

CHAT_ID = os.environ.get("CHAT_ID")
# Канал «🦊 Микита винюхав» (той самий, куди йдуть тендери й документи влади).
# Бюджетні звіти про виконання — публічний факт-привід, тому в канал, не в редакцію.
VYNIUHAV_CHAT_ID = os.environ.get("DOCUMENTS_CHAT_ID") or os.environ.get("PROZORRO_CHAT_ID")

_ALLOWED_USER_IDS = {
    int(uid)
    for uid in os.environ.get("ALLOWED_USER_IDS", "").split(",")
    if uid.strip()
}

# ключ у sync_state нори: JSON-список URL уже завантажених файлів
_SEEN_KEY = "budget_snapshots_seen"

# поріг «розпорядник не виконує»: % касових видатків до плану ЗВІТНОГО періоду
UNDERPERFORM_PCT = 80

_SCHEMA_STATEMENTS = [
    "CREATE SCHEMA IF NOT EXISTS budget",
    # Один снапшот = один xlsx (витрати або надходження) станом на дату
    """
    CREATE TABLE IF NOT EXISTS budget.snapshot (
        id            serial PRIMARY KEY,
        fiscal_year   int NOT NULL,
        snapshot_date date NOT NULL,
        kind          text NOT NULL CHECK (kind IN ('expenditure','revenue')),
        source_file   text,
        loaded_at     timestamptz NOT NULL DEFAULT now(),
        UNIQUE (fiscal_year, snapshot_date, kind)
    )
    """,
    # Витрати: рядки по розпорядниках (code_type='unit', code=КВК) і
    # деталізація під ними по КЕКВ (Лист1) та КБП-функціях (Лист2).
    # 'total' — рядок «Разом».
    """
    CREATE TABLE IF NOT EXISTS budget.snapshot_expenditure_line (
        id          bigserial PRIMARY KEY,
        snapshot_id int NOT NULL REFERENCES budget.snapshot,
        kvk         text NOT NULL,
        unit_name   text,
        code_type   text NOT NULL CHECK (code_type IN ('unit','kekv','kbp','total')),
        code        text NOT NULL,
        line_name   text,
        annual_plan numeric(16,2),
        period_plan numeric(16,2),
        actual      numeric(16,2),
        pct         numeric(9,4),
        UNIQUE (snapshot_id, kvk, code_type, code)
    )
    """,
    # Надходження: кодів у файлі немає — назви показників по фондах,
    # line_order зберігає порядок документа. fund='total' — «Всього доходів».
    """
    CREATE TABLE IF NOT EXISTS budget.snapshot_revenue_line (
        id          bigserial PRIMARY KEY,
        snapshot_id int NOT NULL REFERENCES budget.snapshot,
        fund        text NOT NULL CHECK (fund IN ('general','special','total')),
        line_order  int NOT NULL,
        line_name   text NOT NULL,
        is_total    boolean DEFAULT false,
        annual_plan numeric(16,2),
        period_plan numeric(16,2),
        actual      numeric(16,2),
        deviation   numeric(16,2),
        pct_annual  numeric(9,4),
        pct_period  numeric(9,4),
        UNIQUE (snapshot_id, fund, line_order)
    )
    """,
]

_ensure_done = False


def ensure_snapshot_schema():
    global _ensure_done
    if _ensure_done:
        return
    for sql in _SCHEMA_STATEMENTS:
        bot_db.execute(sql)
    _ensure_done = True


def is_ready():
    return bot_db.is_configured()


# ---------- Парсери ----------

# Дата в імені: на сайті через крапки (02.01.2026), але при збереженні/
# пересиланні крапки й пробіли стають підкресленнями (02_01_2026) — приймаємо
# крапку, підкреслення, дефіс і пробіл як розділювач.
_DATE_RE = re.compile(r"(\d{2})[._\-\s](\d{2})[._\-\s](\d{4})")
_UNIT_RE = re.compile(r"^(\d{2})\s+(\S.*)")
_DETAIL_RE = re.compile(r"^(\d{4})\s+(\S.*)")


def _money(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(round(v, 2)))
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _pct(v):
    """Відсоток: число або None (буває текст «в 1,9 р.б.»)."""
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    return None


def looks_like_snapshot(data):
    """Швидка перевірка «це щомісячна інформація?» для роутингу в
    budget_package_handler (без повного парсингу)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        wb.close()
        title = " ".join(str(v) for v in first if isinstance(v, str)).lower()
        return "щомісячна інформація" in title
    except Exception:  # noqa: BLE001
        return False


def parse_expenditure_snapshot(data):
    """«Витрати щомісячна інформація»: Лист1 = КВК×КЕКВ, Лист2 = КВК×КБП.
    Повертає {'lines': [...], 'total': {...} або None}."""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    lines, total = [], None
    units_seen = set()
    for si, sheet in enumerate(wb.sheetnames):
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        # маркер типу деталізації — підзаголовок під «КВК» («КЕКВ» / «КБП…»)
        marker = " ".join(
            str(c) for row in rows[:6] for c in row if isinstance(c, str)
        ).upper()
        detail_type = "kbp" if "КБП" in marker else "kekv"
        kvk, unit_name = None, None
        for row in rows:
            name = row[0] if row else None
            if not isinstance(name, str) or not name.strip():
                continue
            name = re.sub(r"\s+", " ", name).strip()
            nums = {
                "annual_plan": _money(row[1] if len(row) > 1 else None),
                "period_plan": _money(row[2] if len(row) > 2 else None),
                "actual": _money(row[3] if len(row) > 3 else None),
                "pct": _pct(row[4] if len(row) > 4 else None),
            }
            if name.lower().startswith("разом"):
                # після «Разом» іде деталізація ПО ВСЬОМУ бюджету (КЕКВ/КБП
                # без розпорядника) — кладемо її під синтетичний КВК '00'
                if total is None:
                    total = nums
                kvk, unit_name = "00", "Разом (весь бюджет)"
                continue
            m = _UNIT_RE.match(name)
            if m and len(m.group(1)) == 2:
                kvk, unit_name = m.group(1), m.group(2)
                # рядок розпорядника однаковий на обох листах — пишемо раз
                if kvk not in units_seen:
                    units_seen.add(kvk)
                    lines.append({"kvk": kvk, "unit_name": unit_name,
                                  "code_type": "unit", "code": kvk,
                                  "line_name": unit_name, **nums})
                continue
            m = _DETAIL_RE.match(name)
            if m and kvk:
                lines.append({"kvk": kvk, "unit_name": unit_name,
                              "code_type": detail_type, "code": m.group(1),
                              "line_name": m.group(2), **nums})
    wb.close()
    if not lines:
        raise ValueError("Не знайшов жодного рядка розпорядника (КВК) у файлі витрат")
    return {"lines": lines, "total": total}


def parse_revenue_snapshot(data):
    """«Надходження щомісячна інформація»: показники за назвами по фондах."""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    fund, order, lines = None, 0, []
    for row in ws.iter_rows(values_only=True):
        name = row[0] if row else None
        if not isinstance(name, str) or not name.strip():
            continue
        name = re.sub(r"\s+", " ", name).strip()
        low = name.lower()
        if low == "загальний фонд":
            fund = "general"
            continue
        if low == "спеціальний фонд":
            fund = "special"
            continue
        if low.startswith("щомісячна інформація") or low.startswith("найменування"):
            continue
        this_fund = fund
        is_total = low.startswith("всього доходів")
        if low == "всього доходів":
            this_fund = "total"
        if this_fund is None:
            continue
        order += 1
        lines.append({
            "fund": this_fund, "line_order": order, "line_name": name,
            "is_total": is_total,
            "annual_plan": _money(row[1] if len(row) > 1 else None),
            "period_plan": _money(row[2] if len(row) > 2 else None),
            "actual": _money(row[3] if len(row) > 3 else None),
            "deviation": _money(row[4] if len(row) > 4 else None),
            "pct_annual": _pct(row[5] if len(row) > 5 else None),
            "pct_period": _pct(row[6] if len(row) > 6 else None),
        })
    wb.close()
    if not lines:
        raise ValueError("Не знайшов жодного показника у файлі надходжень")
    return {"lines": lines}


# ---------- Завантаження ----------

def _snapshot_date_from_name(filename):
    m = _DATE_RE.search(filename)
    if not m:
        raise ValueError(
            f"В імені файлу немає дати DD.MM.YYYY: {filename} — у шапках xlsx "
            "її не пишуть, тому дата зрізу береться з імені"
        )
    return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date()


def load_snapshot(data, filename):
    """Вантажить один xlsx щомісячної інформації. Ідемпотентно (повтор —
    перезапис рядків). Синхронна — викликати через asyncio.to_thread."""
    ensure_snapshot_schema()
    snap_date = _snapshot_date_from_name(filename)
    fiscal_year = (snap_date - timedelta(days=1)).year

    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    wb.close()
    title = " ".join(str(v) for v in first if isinstance(v, str)).lower()
    if "надходження" in title:
        kind, parsed = "revenue", parse_revenue_snapshot(data)
    elif "використання коштів" in title or "витрат" in title:
        kind, parsed = "expenditure", parse_expenditure_snapshot(data)
    else:
        raise ValueError(f"Не впізнав тип щомісячної інформації: «{title[:80]}»")

    import psycopg2.extras
    conn = bot_db._connect()
    try:
        with conn, conn.cursor() as cur:
            cur.execute(
                "INSERT INTO budget.snapshot (fiscal_year, snapshot_date, kind, source_file) "
                "VALUES (%s, %s, %s, %s) "
                "ON CONFLICT (fiscal_year, snapshot_date, kind) "
                "DO UPDATE SET source_file = EXCLUDED.source_file, loaded_at = now() "
                "RETURNING id",
                (fiscal_year, snap_date, kind, filename),
            )
            snapshot_id = cur.fetchone()[0]
            if kind == "expenditure":
                cur.execute(
                    "DELETE FROM budget.snapshot_expenditure_line WHERE snapshot_id = %s",
                    (snapshot_id,),
                )
                cols = ["kvk", "unit_name", "code_type", "code", "line_name",
                        "annual_plan", "period_plan", "actual", "pct"]
                # депфін інколи дублює цілі блоки (лютий-2026: загальноміський
                # КЕКВ двічі, значення ідентичні) — перший виграє
                seen_keys, rows = set(), []
                for ln in parsed["lines"]:
                    key = (ln["kvk"], ln["code_type"], ln["code"])
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)
                    rows.append(tuple([snapshot_id] + [ln.get(c) for c in cols]))
                if parsed["total"]:
                    rows.append(tuple([snapshot_id, "", None, "total", "total",
                                       "Разом"] + [parsed["total"][c] for c in
                                                   ("annual_plan", "period_plan", "actual", "pct")]))
                psycopg2.extras.execute_values(
                    cur,
                    f"INSERT INTO budget.snapshot_expenditure_line "
                    f"(snapshot_id, {', '.join(cols)}) VALUES %s",
                    rows, page_size=200,
                )
                n = len(rows)
            else:
                cur.execute(
                    "DELETE FROM budget.snapshot_revenue_line WHERE snapshot_id = %s",
                    (snapshot_id,),
                )
                cols = ["fund", "line_order", "line_name", "is_total", "annual_plan",
                        "period_plan", "actual", "deviation", "pct_annual", "pct_period"]
                rows = [tuple([snapshot_id] + [ln.get(c) for c in cols])
                        for ln in parsed["lines"]]
                psycopg2.extras.execute_values(
                    cur,
                    f"INSERT INTO budget.snapshot_revenue_line "
                    f"(snapshot_id, {', '.join(cols)}) VALUES %s",
                    rows, page_size=200,
                )
                n = len(rows)
    finally:
        conn.close()
    return {"snapshot_id": snapshot_id, "kind": kind, "date": snap_date,
            "fiscal_year": fiscal_year, "lines": n,
            "total": parsed.get("total") if kind == "expenditure" else None}


# ---------- Аналітика ----------

def execution_report(fiscal_year=None):
    """Звіт по останньому снапшоту витрат: розпорядники з найгіршим %
    виконання плану звітного періоду + звірка планів з ревізією рішень.
    Повертає dict або None, якщо снапшотів ще немає."""
    ensure_snapshot_schema()
    where = "WHERE kind = 'expenditure'"
    params = []
    if fiscal_year:
        where += " AND fiscal_year = %s"
        params.append(fiscal_year)
    snap = bot_db.query(
        f"SELECT * FROM budget.snapshot {where} ORDER BY snapshot_date DESC LIMIT 1",
        params or None,
    )
    if not snap:
        return None
    snap = snap[0]
    units = bot_db.query(
        "SELECT kvk, unit_name, annual_plan, period_plan, actual, pct "
        "FROM budget.snapshot_expenditure_line "
        "WHERE snapshot_id = %s AND code_type = 'unit' ORDER BY pct NULLS LAST",
        (snap["id"],),
    )
    total = bot_db.query(
        "SELECT annual_plan, period_plan, actual, pct "
        "FROM budget.snapshot_expenditure_line "
        "WHERE snapshot_id = %s AND code_type = 'total'",
        (snap["id"],),
    )
    # звірка з останньою ревізією рішень (методологія знімка інша — інформативно).
    # Ревізій може ще не бути (пакети рішень не завантажені) — тоді просто без звірки.
    revision, rev_units = None, {}
    try:
        revision = bot_db.query(
            """SELECT r.decision_number, r.effective_order,
                      (SELECT COALESCE(SUM(e.total),0) FROM budget.plan_expenditure_line e
                        WHERE e.revision_id = r.id AND NOT e.is_unit_total) AS total
               FROM budget.plan_revision r
               WHERE r.fiscal_year = %s
                 AND EXISTS (SELECT 1 FROM budget.plan_expenditure_line x WHERE x.revision_id = r.id)
               ORDER BY r.effective_order DESC LIMIT 1""",
            (snap["fiscal_year"],),
        )
        if revision:
            rev_id = bot_db.query(
                "SELECT id FROM budget.plan_revision WHERE fiscal_year = %s AND effective_order = %s",
                (snap["fiscal_year"], revision[0]["effective_order"]),
            )[0]["id"]
            for row in bot_db.query(
                "SELECT kpkvk, total FROM budget.plan_expenditure_line "
                "WHERE revision_id = %s AND is_unit_total AND kpkvk LIKE '%%00000' "
                "AND kpkvk NOT LIKE '%%10000'",
                (rev_id,),
            ):
                rev_units[row["kpkvk"][:2]] = row["total"]
    except Exception:  # noqa: BLE001 — таблиць ревізій ще немає
        revision, rev_units = None, {}
    return {"snapshot": snap, "units": units, "total": total[0] if total else None,
            "revision": revision[0] if revision else None, "rev_units": rev_units}


def _fmt_money(v):
    if v is None:
        return "—"
    return f"{v:,.0f}".replace(",", " ")


def format_execution_message(rep):
    snap = rep["snapshot"]
    lines = [
        f"🦊 Виконання бюджету станом на {snap['snapshot_date'].strftime('%d.%m.%Y')} "
        f"(рік {snap['fiscal_year']}):"
    ]
    if rep["total"]:
        t = rep["total"]
        lines.append(
            f"Разом: план періоду {_fmt_money(t['period_plan'])} грн, "
            f"касові {_fmt_money(t['actual'])} грн — {t['pct']:.1f}%"
            if t["pct"] is not None else
            f"Разом: план періоду {_fmt_money(t['period_plan'])} грн, касові {_fmt_money(t['actual'])} грн"
        )
    laggards = [u for u in rep["units"] if u["pct"] is not None and float(u["pct"]) < UNDERPERFORM_PCT]
    if laggards:
        lines.append(f"\n❗ Виконали план періоду менш як на {UNDERPERFORM_PCT}%:")
        for u in laggards:
            lines.append(f"• {u['kvk']} {u['unit_name']}: {float(u['pct']):.1f}% "
                         f"({_fmt_money(u['actual'])} з {_fmt_money(u['period_plan'])} грн)")
    else:
        lines.append(f"\nУсі розпорядники виконали план періоду на ≥{UNDERPERFORM_PCT}%.")
    if rep["revision"]:
        rev = rep["revision"]
        lines.append(
            f"\n📋 Звірка з рішеннями: остання ревізія {rev['decision_number']} — "
            f"{_fmt_money(rev['total'])} грн; річний план у знімку Казначейства — "
            f"{_fmt_money(rep['total']['annual_plan']) if rep['total'] else '—'} грн."
        )
        lines.append(
            "Різниця = міжсесійні зміни (розпорядження мера/виконком) ± методологія "
            "знімка (без кредитів і власних надходжень установ)."
        )
        # найбільші зсуви плану по розпорядниках
        diffs = []
        for u in rep["units"]:
            rv = rep["rev_units"].get(u["kvk"])
            if rv is not None and u["annual_plan"] is not None:
                d = u["annual_plan"] - rv
                if d != 0:
                    diffs.append((abs(d), u["kvk"], u["unit_name"], d))
        diffs.sort(reverse=True)
        if diffs:
            lines.append("\nНайбільші зсуви річного плану проти останнього рішення:")
            for _, kvk, name, d in diffs[:6]:
                sign = "+" if d > 0 else "−"
                lines.append(f"• {kvk} {name}: {sign}{_fmt_money(abs(d))} грн")
    return "\n".join(lines)


# ---------- Монітор сторінки міськради ----------

def _page_links(year):
    """Список (url, filename) xlsx щомісячної інформації за рік, нові першими
    (порядок як на сторінці)."""
    url = PAGE_URL_TMPL.format(year=year)
    response = requests.get(url, timeout=20)
    if response.status_code != 200:
        raise RuntimeError(f"Сторінка щомісячної інформації: HTTP {response.status_code}")
    # сайт віддає кириличні href без percent-encoding; requests вгадує
    # кодування як latin-1 і лінки перетворюються на кашу → примусово utf-8
    response.encoding = "utf-8"
    links = []
    for href in re.findall(r'href="([^"]+\.(?:xlsx|XLSX))"', response.text):
        full = href if href.startswith("http") else MKRADA_BASE + href
        name = urllib.parse.unquote(full.split("/")[-1])
        links.append((full, name))
    return links


def _get_seen():
    raw = bot_db.get_state(_SEEN_KEY)
    return set(json.loads(raw)) if raw else set()


def _set_seen(seen):
    bot_db.set_state(_SEEN_KEY, json.dumps(sorted(seen), ensure_ascii=False))


async def check_monthly_snapshots(bot, chat_id=None, force_report=False):
    """Щоденна перевірка сторінки міськради. Нові файли → у нору + анонс у канал
    «винюхав». Перший запуск: бекфіл усієї історії року тихо, звіт лише за
    останній місяць (N=1 місяць — перевірити формат відправки без спаму)."""
    if not is_ready():
        return
    chat_id = chat_id or VYNIUHAV_CHAT_ID
    year = datetime.now().year
    try:
        links = await asyncio.to_thread(_page_links, year)
        # у січні нові файли за грудень ще лежать на сторінці минулого року
        if datetime.now().month == 1:
            links += await asyncio.to_thread(_page_links, year - 1)
    except Exception as e:  # noqa: BLE001
        await notify_error(bot, "budget_snapshots", f"сторінка щомісячної інформації: {e}")
        return

    seen = await asyncio.to_thread(_get_seen)
    first_run = not seen
    new = [(u, n) for u, n in links if u not in seen]
    if not new:
        return

    loaded, failed = [], []
    for url, name in new:
        try:
            response = await asyncio.to_thread(requests.get, url, timeout=60)
            response.raise_for_status()
            rep = await asyncio.to_thread(load_snapshot, response.content, name)
            loaded.append(rep)
            seen.add(url)
        except Exception as e:  # noqa: BLE001
            failed.append((name, str(e)))
    await asyncio.to_thread(_set_seen, seen)

    if failed:
        await notify_error(
            bot, "budget_snapshots",
            "не завантажились: " + "; ".join(f"{n} — {e[:80]}" for n, e in failed),
        )
    if not loaded:
        return

    exp_loaded = [r for r in loaded if r["kind"] == "expenditure"]
    latest = max(loaded, key=lambda r: r["date"])["date"]
    if first_run and not force_report:
        # baseline: історію залили тихо, анонс і розбір — лише за останній місяць
        note = (
            "🦊 Микита винюхав свіжу інформацію про те, як наші улюблені "
            "чиновники витрачають народні гроші — місто оновило звіт про "
            f"виконання бюджету станом на {latest.strftime('%d.%m.%Y')}."
        )
        await bot.send_message(chat_id=chat_id, text=note)
        if exp_loaded:
            rep = await asyncio.to_thread(execution_report)
            if rep:
                await bot.send_message(chat_id=chat_id, text=format_execution_message(rep))
        return

    # Один анонс на кожну нову дату (файлів двоє на місяць — витрати й
    # надходження, але привід один)
    for d in sorted({r["date"] for r in loaded}):
        await bot.send_message(
            chat_id=chat_id,
            text=(
                "🦊 Микита винюхав свіжу інформацію про те, як наші улюблені "
                "чиновники витрачають народні гроші — місто оновило звіт про "
                f"виконання бюджету станом на {d.strftime('%d.%m.%Y')}."
            ),
        )
    if exp_loaded:
        rep = await asyncio.to_thread(execution_report)
        if rep:
            await bot.send_message(chat_id=chat_id, text=format_execution_message(rep))


async def run_snapshot_check(bot):
    """Обгортка для APScheduler (щодня; тихо, коли нового немає)."""
    try:
        await check_monthly_snapshots(bot)
    except Exception as e:  # noqa: BLE001
        await notify_error(bot, "budget_snapshots", str(e))


# ---------- Telegram-хендлери ----------

async def budget_execution_handler(update, context):
    """/budget_execution — виконання бюджету по розпорядниках. Пріоритет —
    квартальний офіційний звіт (детальніше, по КПКВК); якщо його немає —
    останній місячний снапшот міськради."""
    msg = update.effective_message
    if not is_ready():
        await msg.reply_text("Нора не налаштована (BOT_DATABASE_URL).")
        return
    # 1) квартальний звіт про виконання, якщо є
    try:
        from handlers import budget_execution_report as ber
        q = await asyncio.to_thread(ber.latest_report)
        if q:
            text = await asyncio.to_thread(ber.format_report, q)
            await msg.reply_text(text, parse_mode="HTML")
            return
    except Exception as e:  # noqa: BLE001
        await msg.reply_text(f"❌ {e}")
        return
    # 2) інакше — місячний снапшот
    try:
        rep = await asyncio.to_thread(execution_report)
    except Exception as e:  # noqa: BLE001
        await msg.reply_text(f"❌ {e}")
        return
    if not rep:
        await msg.reply_text(
            "Даних про виконання ще немає. Кинь у приват квартальний «Звіт про "
            "виконання…» або запусти /budget_snapshot_check для місячних знімків."
        )
        return
    await msg.reply_text(format_execution_message(rep))


async def budget_execution_test_handler(update, context):
    """/budget_execution_test — надіслати зразок анонсу «Микита винюхав…» +
    розбір виконання у канал «винюхав» (в обхід «seen», щоб перевірити формат
    без нового файлу на сайті). За зразком /documents_test, /builder_test."""
    msg = update.effective_message
    if _ALLOWED_USER_IDS and update.effective_user.id not in _ALLOWED_USER_IDS:
        await msg.reply_text("⛔ Тільки для редакції.")
        return
    if not is_ready():
        await msg.reply_text("Нора не налаштована (BOT_DATABASE_URL).")
        return
    rep = await asyncio.to_thread(execution_report)
    if not rep:
        await msg.reply_text(
            "У норі ще немає місячних знімків — спочатку /budget_snapshot_check."
        )
        return
    d = rep["snapshot"]["snapshot_date"]
    note = (
        "🦊 Микита винюхав свіжу інформацію про те, як наші улюблені чиновники "
        "витрачають народні гроші — місто оновило звіт про виконання бюджету "
        f"станом на {d.strftime('%d.%m.%Y')}."
    )
    await context.bot.send_message(chat_id=VYNIUHAV_CHAT_ID, text=note)
    await context.bot.send_message(chat_id=VYNIUHAV_CHAT_ID, text=format_execution_message(rep))
    await msg.reply_text("✅ Зразок анонсу + розбір надіслано в канал «винюхав».")


async def budget_snapshot_check_handler(update, context):
    """/budget_snapshot_check — перевірити сторінку міськради зараз."""
    msg = update.effective_message
    if _ALLOWED_USER_IDS and update.effective_user.id not in _ALLOWED_USER_IDS:
        await msg.reply_text("⛔ Тільки для редакції.")
        return
    if not is_ready():
        await msg.reply_text("Нора не налаштована (BOT_DATABASE_URL).")
        return
    before = len(await asyncio.to_thread(_get_seen)) if bot_db.is_configured() else 0
    await msg.reply_text("🦊 Перевіряю сторінку щомісячної інформації…")
    await check_monthly_snapshots(context.bot, chat_id=update.effective_chat.id)
    after = len(await asyncio.to_thread(_get_seen))
    if after == before:
        await msg.reply_text("Нових файлів немає.")


async def load_snapshot_from_message(msg, context, data, filename):
    """Виклик з budget_package_handler: файл «Щомісячна інформація» у приваті."""
    progress = await msg.reply_text(f"🦊 Це щомісячний снапшот — розбираю {filename}…")
    try:
        rep = await asyncio.to_thread(load_snapshot, data, filename)
    except Exception as e:  # noqa: BLE001
        await progress.edit_text(f"❌ Не завантажилось: {e}")
        return
    kind_ua = "витрати" if rep["kind"] == "expenditure" else "надходження"
    text = (f"✅ Снапшот: {kind_ua} станом на {rep['date'].strftime('%d.%m.%Y')} "
            f"(рік {rep['fiscal_year']}), {rep['lines']} рядків.")
    await progress.edit_text(text)
    if rep["kind"] == "expenditure":
        full = await asyncio.to_thread(execution_report)
        if full and full["snapshot"]["id"] == rep["snapshot_id"]:
            await msg.reply_text(format_execution_message(full))
