"""
Квартальні звіти про виконання бюджету Миколаєва (схема budget).

Джерело — офіційний пакет «Звіт про виконання бюджету … за N квартал РРРР»
(казначейські форми, ~18 xlsx). На відміну від місячної «Щомісячної
інформації» (budget_snapshots.py — розріз КВК×КЕКВ), квартальний звіт дає
виконання ПО КОЖНІЙ ПРОГРАМІ (КПКВК 7 знаків) і стикується з ревізіями
рішень напряму по КПКВК.

Вантажимо два найцінніші файли пакета, решту (борг, гарантії, резервний
фонд, заборгованість, трансферти) — пропускаємо:
- «Додаток 13 … розподілу видатків» — видатки по КПКВК, три підрядки на
  програму: «затверджено місцевою радою на звітний рік», «…з урахуванням
  змін» (план з поправками), «виконано за звітний період». 16 колонок як у
  Додатку 3 (загальний/спеціальний фонд + разом);
- «(Форма №2кмб Доходи)» — доходи по коду (8 знаків): затверджено розписом
  і виконано, по загальному (doh_zf) і спеціальному (doh_sf) фондах.

Структура файлів (перевірено на пакеті за I квартал 2026):
- шапка з «Звіт про виконання … за N квартал РРРР року»; повторювані
  колонтитули «АС Є-ЗВІТНІСТЬ / ст. X з Y» і рядок нумерації 1..16 всередині
  таблиці — пропускаємо;
- рядок-заголовок програми: КПКВК у кол.1, у грошових кол. — «Х»; далі
  підрядки з мітками у кол.4;
- підсумки розпорядників — КПКВК виду XX00000/XX10000 без ТПКВК (is_unit_total).

Період і рік — з назви пакета/файлів («за 1 квартал 2026» → year=2026, Q1).

Команди/UX: кинь ZIP пакета в приват — бот сам розпізнає (роутинг з
budget_revisions.budget_package_handler за назвою «Звіт про виконання»);
/budget_execution показує останній наявний зріз (квартальний, якщо є, інакше
місячний снапшот). Тихо не працює без BOT_DATABASE_URL.
"""

import re
import zipfile
from decimal import Decimal, InvalidOperation
from io import BytesIO

from handlers import bot_db

_SCHEMA_STATEMENTS = [
    "CREATE SCHEMA IF NOT EXISTS budget",
    """
    CREATE TABLE IF NOT EXISTS budget.execution_report (
        id          serial PRIMARY KEY,
        fiscal_year int NOT NULL,
        period      text NOT NULL,          -- 'Q1'..'Q4'
        source_file text,
        loaded_at   timestamptz NOT NULL DEFAULT now(),
        UNIQUE (fiscal_year, period)
    )
    """,
    # Видатки по КПКВК: план (з урахуванням змін), первинно затверджений, виконано
    """
    CREATE TABLE IF NOT EXISTS budget.execution_expenditure_line (
        id                  bigserial PRIMARY KEY,
        report_id           int NOT NULL REFERENCES budget.execution_report,
        kpkvk               text NOT NULL,
        tpkvk               text,
        kfk                 text,
        line_name           text,
        is_unit_total       boolean DEFAULT false,
        plan_approved       numeric(16,2),   -- затверджено радою на рік
        plan_amended        numeric(16,2),   -- затверджено з урахуванням змін
        executed            numeric(16,2),   -- виконано за період
        plan_amended_general numeric(16,2),
        executed_general     numeric(16,2),
        plan_amended_special numeric(16,2),
        executed_special     numeric(16,2),
        UNIQUE (report_id, kpkvk)
    )
    """,
    # Доходи по коду
    """
    CREATE TABLE IF NOT EXISTS budget.execution_revenue_line (
        id          bigserial PRIMARY KEY,
        report_id   int NOT NULL REFERENCES budget.execution_report,
        fund        text NOT NULL CHECK (fund IN ('general','special')),
        code        text NOT NULL,
        line_name   text,
        plan_annual numeric(16,2),
        executed    numeric(16,2),
        UNIQUE (report_id, fund, code)
    )
    """,
]

_ensure_done = False


def ensure_execution_schema():
    global _ensure_done
    if _ensure_done:
        return
    for sql in _SCHEMA_STATEMENTS:
        bot_db.execute(sql)
    _ensure_done = True


def is_ready():
    return bot_db.is_configured()


# ---------- Парсинг ----------

_QUARTER_RE = re.compile(r"за\s+([1-4IV]+)\s+квартал\s+(20\d{2})", re.I)
_ROMAN = {"I": 1, "II": 2, "III": 3, "IV": 4}


def _period_year(text):
    """'за 1 квартал 2026' / 'за I квартал 2026 року' → ('Q1', 2026)."""
    m = _QUARTER_RE.search(text or "")
    if not m:
        return None, None
    q = m.group(1).upper()
    n = _ROMAN.get(q) or (int(q) if q.isdigit() else None)
    if not n:
        return None, None
    return f"Q{n}", int(m.group(2))


def _money(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(round(v, 2)))
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s or s.upper() in ("Х", "X", "-"):
        return None
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _code(v, width):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return str(int(round(v))).zfill(width)
    s = str(v).strip()
    if re.fullmatch(r"\d+(\.0+)?", s):
        return str(int(float(s))).zfill(width)
    return None


def looks_like_execution_report(data=None, filename="", names=None):
    """Розпізнавання квартального пакета/файла — за назвою (пакет) або шапкою."""
    hay = (filename or "") + " " + " ".join(names or [])
    if "звіт про виконання" in hay.lower() and "квартал" in hay.lower():
        return True
    if data is not None:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            head = " ".join(
                str(v) for row in ws.iter_rows(min_row=1, max_row=6, values_only=True)
                for v in row if isinstance(v, str)
            ).lower()
            wb.close()
            return "звіт про виконання" in head and "квартал" in head
        except Exception:  # noqa: BLE001
            return False
    return False


def parse_dodatok13(data):
    """«Додаток 13 … розподілу видатків» → список рядків по КПКВК з планом і
    виконанням. Три підрядки на програму (затв. / затв. зі змінами / виконано)."""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    lines, cur = [], None
    for row in ws.iter_rows(values_only=True):
        c1 = row[0] if len(row) > 0 else None
        c1s = str(c1).strip() if c1 is not None else ""
        kpkvk = _code(c1, 7) if isinstance(c1, (int, float)) or re.fullmatch(r"\d{7}", c1s) else None
        raw4 = row[3] if len(row) > 3 else None
        name4 = re.sub(r"\s+", " ", str(raw4)).strip() if isinstance(raw4, str) else ""
        label = name4.lower()
        if kpkvk:
            tpkvk = _code(row[1] if len(row) > 1 else None, 4)
            cur = {
                "kpkvk": kpkvk, "tpkvk": tpkvk,
                "kfk": _code(row[2] if len(row) > 2 else None, 4),
                "line_name": name4 or None,  # назва — з рядка-заголовка (кол.4)
                "is_unit_total": tpkvk is None and kpkvk[-5:] in ("00000", "10000"),
                "plan_approved": None, "plan_amended": None, "executed": None,
                "plan_amended_general": None, "executed_general": None,
                "plan_amended_special": None, "executed_special": None,
            }
            lines.append(cur)
            continue
        # Фінальний блок «Усього:» (кол.1='Х') та будь-який підсумковий заголовок
        # НЕ належать попередній програмі — інакше грандтотал (6 млрд) прилипає
        # до останнього рядка. Скидаємо cur, підрядки тоталу пропускаються.
        if c1s in ("Х", "X") or re.match(r"^(усього|разом|всього)\b", label):
            cur = None
            continue
        if cur is None or not label:
            continue
        gen = _money(row[4] if len(row) > 4 else None)
        spec = _money(row[9] if len(row) > 9 else None)
        total = _money(row[15] if len(row) > 15 else None)
        # ім'я програми — з підрядка «затверджено … на звітний рік» (кол.4 містить
        # мітку, справжня назва — у рядку-заголовку кол.4, але там «Х»; беремо
        # мітку розпорядника з заголовка окремо нижче не потрібно — назва
        # програми не критична для %, лишаємо КПКВК). Тому name — з мітки лише
        # якщо це заголовок; тут пропускаємо.
        if "з урахуванням змін" in label:
            cur["plan_amended"] = total
            cur["plan_amended_general"] = gen
            cur["plan_amended_special"] = spec
        elif "виконано" in label:
            cur["executed"] = total
            cur["executed_general"] = gen
            cur["executed_special"] = spec
        elif "затверджено" in label:
            cur["plan_approved"] = total
    wb.close()
    # відкидаємо порожні (без плану й виконання)
    return [l for l in lines if l["plan_amended"] is not None or l["executed"] is not None]


def parse_dodatok13_names(data):
    """Назви програм з файлу розподілу: рядок-заголовок кол.4 містить назву
    (не «Х»)? Ні — там «Х». Назва в мітках. Тому окремо не тягнемо; імена
    беремо з ревізій при звіті. Заглушка для сумісності."""
    return {}


def parse_revenue_form(data):
    """«Форма №2кмб Доходи» → доходи по коду, обидва фонди (doh_zf/doh_sf)."""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    out = []
    for sheet in wb.sheetnames:
        fund = "general" if sheet.endswith("zf") else ("special" if sheet.endswith("sf") else None)
        if not fund:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            name = row[0] if len(row) > 0 else None
            code = _code(row[2] if len(row) > 2 else None, 8)
            if not code or not isinstance(name, str):
                continue
            out.append({
                "fund": fund, "code": code,
                "line_name": re.sub(r"\s+", " ", name).strip(),
                "plan_annual": _money(row[3] if len(row) > 3 else None),
                "executed": _money(row[4] if len(row) > 4 else None),
            })
    wb.close()
    return out


# ---------- Завантаження ----------

def _zip_entries(data):
    entries = []
    with zipfile.ZipFile(BytesIO(data)) as z:
        for info in z.infolist():
            if info.is_dir() or info.filename.startswith("__MACOSX"):
                continue
            name = info.filename
            if not (info.flag_bits & 0x800):
                try:
                    name = name.encode("cp437").decode("utf-8")
                except (UnicodeDecodeError, UnicodeEncodeError):
                    pass
            entries.append((name, z.read(info)))
    return entries


def load_execution_package(data, filename):
    """Вантажить квартальний пакет (ZIP) або одиночний xlsx. Ідемпотентно.
    Повертає dict-звіт. Синхронна — через asyncio.to_thread."""
    ensure_execution_schema()
    entries = _zip_entries(data) if filename.lower().endswith(".zip") else [(filename, data)]

    period, year = _period_year(filename)
    if not year:
        for name, _ in entries:
            period, year = _period_year(name)
            if year:
                break
    exp_file = rev_file = None
    for name, blob in entries:
        low = name.lower()
        if not low.endswith(".xlsx"):
            continue
        if "розподілу видатків" in low or "додаток 13" in low:
            exp_file = (name, blob)
        elif "доходи" in low and "2кмб" in low:
            rev_file = (name, blob)
    if not year and exp_file:
        # рік/період з шапки файлу видатків
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(exp_file[1]), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        head = " ".join(str(v) for r in ws.iter_rows(min_row=1, max_row=6, values_only=True)
                        for v in r if isinstance(v, str))
        wb.close()
        period, year = _period_year(head)
    if not year:
        raise ValueError("Не визначив рік/квартал зі звіту про виконання")
    if not exp_file:
        raise ValueError("У пакеті немає файлу «розподілу видатків» (Додаток 13)")

    exp_lines = parse_dodatok13(exp_file[1])
    rev_lines = parse_revenue_form(rev_file[1]) if rev_file else []

    import psycopg2.extras
    conn = bot_db._connect()
    try:
        with conn, conn.cursor() as cur:
            cur.execute(
                "INSERT INTO budget.execution_report (fiscal_year, period, source_file) "
                "VALUES (%s, %s, %s) ON CONFLICT (fiscal_year, period) "
                "DO UPDATE SET source_file = EXCLUDED.source_file, loaded_at = now() RETURNING id",
                (year, period, filename),
            )
            report_id = cur.fetchone()[0]
            cur.execute("DELETE FROM budget.execution_expenditure_line WHERE report_id = %s", (report_id,))
            ecols = ["kpkvk", "tpkvk", "kfk", "line_name", "is_unit_total",
                     "plan_approved", "plan_amended", "executed",
                     "plan_amended_general", "executed_general",
                     "plan_amended_special", "executed_special"]
            seen = set()
            erows = []
            for l in exp_lines:
                if l["kpkvk"] in seen:
                    continue
                seen.add(l["kpkvk"])
                erows.append(tuple([report_id] + [l.get(c) for c in ecols]))
            if erows:
                psycopg2.extras.execute_values(
                    cur,
                    f"INSERT INTO budget.execution_expenditure_line (report_id, {', '.join(ecols)}) VALUES %s",
                    erows, page_size=200,
                )
            cur.execute("DELETE FROM budget.execution_revenue_line WHERE report_id = %s", (report_id,))
            rcols = ["fund", "code", "line_name", "plan_annual", "executed"]
            seenr = set()
            rrows = []
            for l in rev_lines:
                k = (l["fund"], l["code"])
                if k in seenr:
                    continue
                seenr.add(k)
                rrows.append(tuple([report_id] + [l.get(c) for c in rcols]))
            if rrows:
                psycopg2.extras.execute_values(
                    cur,
                    f"INSERT INTO budget.execution_revenue_line (report_id, {', '.join(rcols)}) VALUES %s",
                    rrows, page_size=200,
                )
    finally:
        conn.close()
    return {"report_id": report_id, "year": year, "period": period,
            "exp_lines": len(erows), "rev_lines": len(rrows),
            "has_revenue": bool(rev_file),
            "exp_file": exp_file[0], "rev_file": rev_file[0] if rev_file else None}


# ---------- Аналітика ----------

def latest_report(fiscal_year=None):
    ensure_execution_schema()
    where = "WHERE fiscal_year = %s" if fiscal_year else ""
    params = (fiscal_year,) if fiscal_year else None
    rows = bot_db.query(
        f"SELECT * FROM budget.execution_report {where} "
        "ORDER BY fiscal_year DESC, period DESC LIMIT 1", params)
    return rows[0] if rows else None


def execution_by_unit(report_id):
    """Виконання по розпорядниках (підсумки XX00000) — % виконано/план."""
    return bot_db.query(
        "SELECT kpkvk, plan_amended, executed, "
        "  CASE WHEN plan_amended > 0 THEN round(100*executed/plan_amended, 1) END AS pct "
        "FROM budget.execution_expenditure_line "
        "WHERE report_id = %s AND is_unit_total AND kpkvk LIKE '%%00000' "
        "  AND kpkvk NOT LIKE '%%10000' "
        "ORDER BY pct NULLS LAST",
        (report_id,))


def under_executed_programs(report_id, limit=8, min_plan=1_000_000):
    """Найбільші за планом програми з найгіршим % виконання (не підсумки)."""
    return bot_db.query(
        "SELECT kpkvk, plan_amended, executed, "
        "  CASE WHEN plan_amended > 0 THEN round(100*executed/plan_amended,1) END AS pct "
        "FROM budget.execution_expenditure_line "
        "WHERE report_id = %s AND NOT is_unit_total AND plan_amended >= %s "
        "ORDER BY pct NULLS LAST, plan_amended DESC LIMIT %s",
        (report_id, min_plan, limit))


# ---------- Звіт у чат ----------

def _fmt(v):
    return "—" if v is None else f"{v:,.0f}".replace(",", " ")


def _unit_names(fiscal_year):
    """Назви розпорядників за 2-значним КВК. Джерела: ревізії (XX00000 =
    назва департаменту) як основа, місячні снапшоти — доповнюють."""
    names = {}
    try:
        for r in bot_db.query(
            "SELECT DISTINCT ON (left(l.kpkvk,2)) left(l.kpkvk,2) kvk, l.line_name "
            "FROM budget.plan_expenditure_line l JOIN budget.plan_revision r ON r.id=l.revision_id "
            "WHERE r.fiscal_year=%s AND l.is_unit_total AND l.kpkvk LIKE '%%00000' "
            "  AND l.kpkvk NOT LIKE '%%10000' "
            "ORDER BY left(l.kpkvk,2), r.effective_order DESC",
            (fiscal_year,)):
            if r["line_name"]:
                names[r["kvk"]] = r["line_name"]
    except Exception:  # noqa: BLE001
        pass
    try:
        for r in bot_db.query(
            "SELECT DISTINCT ON (l.kvk) l.kvk, l.unit_name FROM budget.snapshot_expenditure_line l "
            "JOIN budget.snapshot s ON s.id = l.snapshot_id "
            "WHERE s.fiscal_year=%s AND l.code_type='unit' ORDER BY l.kvk, s.snapshot_date DESC",
            (fiscal_year,)):
            names.setdefault(r["kvk"], r["unit_name"])
    except Exception:  # noqa: BLE001
        pass
    return names


def _program_names(fiscal_year):
    """Назви програм по КПКВК з останньої ревізії року."""
    names = {}
    try:
        for r in bot_db.query(
            "SELECT DISTINCT ON (l.kpkvk) l.kpkvk, l.line_name FROM budget.plan_expenditure_line l "
            "JOIN budget.plan_revision r ON r.id=l.revision_id "
            "WHERE r.fiscal_year=%s ORDER BY l.kpkvk, r.effective_order DESC",
            (fiscal_year,)):
            names[r["kpkvk"]] = r["line_name"]
    except Exception:  # noqa: BLE001
        pass
    return names


_PERIOD_UA = {"Q1": "I квартал", "Q2": "I півріччя (II квартал)",
              "Q3": "9 місяців (III квартал)", "Q4": "рік (IV квартал)"}


def format_report(report, under_pct=80):
    """Людський звіт по квартальному виконанню: розпорядники з найгіршим %,
    найбільші недоосвоєні програми."""
    year, period, rid = report["fiscal_year"], report["period"], report["id"]
    units = execution_by_unit(rid)
    unames = _unit_names(year)
    pnames = _program_names(year)
    tot = bot_db.query(
        "SELECT sum(plan_amended) p, sum(executed) e FROM budget.execution_expenditure_line "
        "WHERE report_id=%s AND is_unit_total AND kpkvk LIKE '%%00000' AND kpkvk NOT LIKE '%%10000'",
        (rid,))[0]
    lines = [f"🦊 Виконання бюджету за {_PERIOD_UA.get(period, period)} {year}:"]
    if tot["p"]:
        pct = 100 * tot["e"] / tot["p"] if tot["p"] else 0
        lines.append(f"Разом видатки: план {_fmt(tot['p'])} → виконано {_fmt(tot['e'])} грн — {pct:.1f}%")
    laggards = [u for u in units if u["pct"] is not None and float(u["pct"]) < under_pct]
    if laggards:
        lines.append(f"\n❗ Розпорядники з виконанням &lt;{under_pct}%:")
        for u in laggards:
            nm = unames.get(u["kpkvk"][:2], u["kpkvk"])
            lines.append(f"• {nm}: {float(u['pct']):.1f}% ({_fmt(u['executed'])} з {_fmt(u['plan_amended'])} грн)")
    worst = under_executed_programs(rid, limit=8, min_plan=10_000_000)
    if worst:
        lines.append("\n🔻 Найбільші недоосвоєні програми (план ≥10 млн):")
        for p in worst:
            nm = pnames.get(p["kpkvk"]) or p["kpkvk"]
            pct = f"{float(p['pct']):.1f}%" if p["pct"] is not None else "0%"
            lines.append(f"• {p['kpkvk']} {nm}: {pct} ({_fmt(p['executed'])} з {_fmt(p['plan_amended'])} грн)")
    return "\n".join(lines)


async def load_from_message(msg, context, data, filename):
    """Виклик з budget_package_handler: квартальний пакет/файл у приваті."""
    import asyncio
    progress = await msg.reply_text(f"🦊 Це квартальний звіт про виконання — розбираю {filename}…")
    try:
        r = await asyncio.to_thread(load_execution_package, data, filename)
    except Exception as e:  # noqa: BLE001
        await progress.edit_text(f"❌ Не завантажилось: {e}")
        return
    rev = " + доходи" if r["has_revenue"] else ""
    await progress.edit_text(
        f"✅ Виконання за {_PERIOD_UA.get(r['period'], r['period'])} {r['year']}: "
        f"{r['exp_lines']} програм{rev} у норі."
    )
    report = await asyncio.to_thread(latest_report, r["year"])
    if report and report["id"] == r["report_id"]:
        await msg.reply_text(format_report(report), parse_mode="HTML")
